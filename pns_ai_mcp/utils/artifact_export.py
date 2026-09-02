# -*- coding: utf-8 -*-
# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025-2026 Patanegra Soft <https://patanegra.com>
"""Deterministic file export from tabular rows (no LLM, no domain literals).

When the user names a known download format, serialize public columns to
bytes. Word is Office HTML (``.doc``). PowerPoint stays refused.
Host-testable without Odoo.
"""
from __future__ import annotations

import base64
import html as html_lib
import importlib.util
import io
import logging
import re
import zipfile
from pathlib import Path

_logger = logging.getLogger(__name__)

DOWNLOAD_CTA = 'Download'

_XLSX_MIME = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

# Chatboo client assembles these (table + ECharts PNG). Excel/md stay here
# (icon Excel uses this same writer).
CLIENT_ASSEMBLED_KINDS = frozenset(('doc', 'pdf', 'html'))

# Canonical kind → extension + mimetype. excel/xls still emit xlsx.
FORMAT_SPECS = {
    'pdf': {'kind': 'pdf', 'ext': 'pdf', 'mimetype': 'application/pdf'},
    'excel': {'kind': 'xlsx', 'ext': 'xlsx', 'mimetype': _XLSX_MIME},
    'xlsx': {'kind': 'xlsx', 'ext': 'xlsx', 'mimetype': _XLSX_MIME},
    'xls': {'kind': 'xlsx', 'ext': 'xlsx', 'mimetype': _XLSX_MIME},
    'txt': {'kind': 'txt', 'ext': 'txt', 'mimetype': 'text/plain; charset=utf-8'},
    'md': {'kind': 'md', 'ext': 'md', 'mimetype': 'text/markdown; charset=utf-8'},
    'markdown': {'kind': 'md', 'ext': 'md', 'mimetype': 'text/markdown; charset=utf-8'},
    'html': {'kind': 'html', 'ext': 'html', 'mimetype': 'text/html; charset=utf-8'},
    'word': {'kind': 'doc', 'ext': 'doc', 'mimetype': 'application/msword'},
    'docx': {'kind': 'doc', 'ext': 'doc', 'mimetype': 'application/msword'},
    'doc': {'kind': 'doc', 'ext': 'doc', 'mimetype': 'application/msword'},
}

# ASCII edges only: ``\\b`` misses CJK (``PDFを`` has no Unicode word break).
# A leading dot (``.odt``) is not ``[a-z0-9_]``, so these already match.
_FORMAT_RE = re.compile(
    r'(?i)(?<![a-z0-9_])(?:markdown|xlsx|xls|excel|pdf|html|txt|md|'
    r'docx|word|doc)(?![a-z0-9_])',
)
_REFUSE_RE = re.compile(
    r'(?i)(?<![a-z0-9_])'
    r'(?:pptx|powerpoint|odt|ods|odp|rtf|ppt)'
    r'(?![a-z0-9_])',
)
# Dotted token not in the closed list (``.rtf``, ``.csv``). ``foo.com`` does
# not match: the letter before the dot fails the lookbehind.
_DOT_EXT_RE = re.compile(
    r'(?i)(?<![a-z0-9_])\.([a-z][a-z0-9]{1,4})(?![a-z0-9_])',
)

# Engine injects this when the utterance names a download format. Protocol,
# not a locale verb list (same pattern as ``self``).
SESSION_FILE_PROTOCOL = (
    'The user named a download format. Query structured rows (return '
    'data) and set result __title__ to a short Locale phrase that names '
    'the content (not a copy of the user prompt). Optional: '
    '__filename_stem__ (slug, no extension), __sheet_caption__ (one '
    'line on a single Excel sheet), __column_labels__ (key to visible '
    'header). The server serializes the file and pins a session '
    'download card (same box as a clipboard, a manual attach, or a '
    'Sesame file). The user may name one or more formats; the server '
    'pins one card per format. FORBIDDEN this turn: reportlab, '
    'openpyxl, ir.attachment, and propose_safe_operations for files. '
    'Do not paint a listing unless __show_table__ or show_mode '
    'show-table / chart-table / table. show-chart without '
    '__show_table__ keeps the chart and hides the listing. Word, PDF, '
    'and HTML mirror this turn (table, chart, or both) and are '
    'assembled by the Chatboo client (same file for the session card '
    'and the bubble icons; the canvas paints the chart). Do not '
    'rebuild those files in RelaxAICode. Word is a session download '
    '(Office HTML .doc). Optional __rich_doc__ true wraps this turn '
    'HTML in that Word file (Locale like self). Page flag is '
    '__landscape__ true or false; do not invent other orientation keys. '
    'Excel and Markdown stay server tables (no chart).'
)

SESSION_FILE_META_PROTOCOL = (
    'The rows are ready. Return the SAME data and set __title__ '
    '(Locale: what the file contains). Optional __filename_stem__, '
    '__sheet_caption__, __column_labels__. Word, PDF, and HTML with a '
    'chart are assembled by the Chatboo client. Do not rebuild the '
    'file in Python. Do not call propose_safe_operations.'
)

SESSION_FILE_PROPOSE_ERROR = (
    'Session files are not caja B. Return structured rows (data); '
    'the server will serialize and attach a session download chip. '
    'Do not call propose_safe_operations for ir.attachment.'
)

SESSION_FILE_ATTACHED_NOTE = (
    'The server attached the session download card. Do not propose '
    'ir.attachment. Point the user to the card in one sentence.'
)

# Report / painter-free: named format still keeps the narrative in the bubble.
SESSION_FILE_REPORT_NOTE = (
    'The user also named a download format. Write the FULL report in the '
    'bubble (every outline heading). The server attaches the session '
    'download card under the prose. Do not replace the narrative with a '
    'pointer to the card. Do not call propose_safe_operations for files.'
)

REFUSED_EXPORT_REPLY = (
    'That file type is not available as a session download. '
    'I can send PDF, Excel, Word, text, Markdown, or HTML.'
)

REFUSED_PARTIAL_REPLY = (
    'These file types are not available as a session download: %s. '
    'I can send PDF, Excel, Word, text, Markdown, or HTML.'
)

_CHATBOO_SESSION_MODEL = 'chatboo.session'
_FILE_VALUE_KEYS = ('datas', 'datas_fname', 'db_datas')


def _spec_for_ext(ext):
    """FORMAT_SPECS row for a dotted extension, or None."""
    ext = (ext or '').lower()
    if ext in FORMAT_SPECS:
        return ext, FORMAT_SPECS[ext]
    for key, spec in FORMAT_SPECS.items():
        if spec.get('ext') == ext:
            return key, spec
    return None, None


def _hit_from_token(token):
    """FORMAT_SPECS row for a word token, or None."""
    token = (token or '').lower()
    spec = FORMAT_SPECS.get(token)
    if not spec:
        return None
    return {
        'format': token,
        'refused': False,
        'kind': spec['kind'],
        'ext': spec['ext'],
        'mimetype': spec['mimetype'],
    }


def turn_export_utterance(env, fallback=''):
    """User-typed text for format detection. Never engine-rewritten handoff.

    Skill bodies and report handoffs mention HTML/Markdown as *medium*;
    those words must not spawn download cards.
    """
    ctx = getattr(env, 'context', None) or {}
    if isinstance(ctx, dict):
        raw = ctx.get('user_message')
        if isinstance(raw, str) and raw.strip():
            return raw
    if isinstance(fallback, str):
        return fallback
    return ''


def payload_has_report_outline(payload):
    """True when the skill declared a narrative skeleton (keep prose)."""
    if not isinstance(payload, dict):
        return False
    outline = payload.get('report_outline')
    if not isinstance(outline, (list, tuple)):
        return False
    return any(isinstance(h, str) and h.strip() for h in outline)


def parse_requested_export(user_message):
    """Return format intent from the utterance, or None.

    File-type tokens only (pdf/xlsx/…), not a locale verb list. Table vs
    card is the always-on pack (same pattern as ``self``). Collects every
    unique serializable kind in appearance order. Known formats win over
    a refused token in the same sentence: ``refused`` is True only when
    nothing valid was named. Compat: ``format`` / ``kind`` are the first
    valid hit; ``formats`` and ``refused_tokens`` list the rest.
    """
    text = user_message if isinstance(user_message, str) else ''
    if not text.strip():
        return None
    events = []
    for match in _FORMAT_RE.finditer(text):
        hit = _hit_from_token(match.group(0))
        if hit:
            events.append((match.start(), 'ok', hit, match.group(0)))
    for match in _DOT_EXT_RE.finditer(text):
        ext = match.group(1)
        token, spec = _spec_for_ext(ext)
        if spec:
            events.append((match.start(), 'ok', {
                'format': token,
                'refused': False,
                'kind': spec['kind'],
                'ext': spec['ext'],
                'mimetype': spec['mimetype'],
            }, match.group(0)))
        else:
            events.append((match.start(), 'refuse', None, match.group(0)))
    for match in _REFUSE_RE.finditer(text):
        events.append((match.start(), 'refuse', None, match.group(0)))
    if not events:
        return None
    events.sort(key=lambda item: item[0])
    formats = []
    seen_kinds = set()
    refused_tokens = []
    seen_refuse = set()
    for _pos, etype, hit, raw in events:
        if etype == 'ok':
            kind = hit['kind']
            if kind in seen_kinds:
                continue
            seen_kinds.add(kind)
            formats.append(hit)
            continue
        label = ' '.join(str(raw or '').split())
        if not label:
            continue
        key = label.lower().lstrip('.')
        if not key or key in seen_refuse:
            continue
        seen_refuse.add(key)
        refused_tokens.append(label)
    if formats:
        first = dict(formats[0])
        first['formats'] = formats
        first['refused_tokens'] = refused_tokens
        first['refused'] = False
        return first
    return {
        'format': None,
        'refused': True,
        'formats': [],
        'refused_tokens': refused_tokens,
    }


# Shape heuristic: many columns (or a very wide header row) need A4 landscape.
PDF_LANDSCAPE_MIN_COLS = 6
PDF_LANDSCAPE_MIN_HEADER_CHARS = 48


def pdf_column_count(payload=None, rows=None):
    """Public column count for the landscape heuristic, or None if unknown."""
    public = public_rows(tabular_rows(payload, rows) or [])
    if not public:
        return None
    return len(public_columns(public))


def _payload_landscape_flag(payload):
    """``__landscape__`` wins; ``__orientation__`` is landscape|portrait only."""
    if not isinstance(payload, dict):
        return None
    if '__landscape__' in payload:
        return bool(payload.get('__landscape__'))
    raw = payload.get('__orientation__')
    if raw is None or isinstance(raw, bool):
        return None
    token = str(raw).strip().lower()
    if token == 'landscape':
        return True
    if token == 'portrait':
        return False
    return None


def pdf_landscape(payload=None, default=None, col_count=None, rows=None):
    """Session PDF orientation.

    ``__landscape__`` on the payload always wins (pack / Locale, like
    ``self``). ``__orientation__`` accepts only the product tokens
    landscape / portrait. Otherwise a shape heuristic: wide tables go
    landscape. Not a locale word list.
    """
    flagged = _payload_landscape_flag(payload)
    if flagged is not None:
        return flagged
    n = col_count
    keys = None
    if n is None:
        public = public_rows(tabular_rows(payload, rows) or [])
        if public:
            keys = public_columns(public)
            n = len(keys)
    if n is not None:
        if n >= PDF_LANDSCAPE_MIN_COLS:
            return True
        if keys is None and rows:
            keys = public_columns(public_rows(rows))
        if keys and sum(len(str(k)) for k in keys) >= PDF_LANDSCAPE_MIN_HEADER_CHARS:
            return True
        return False
    return True if default is None else bool(default)


_CHART_ONLY_SHOW_MODES = frozenset(('show-chart', 'chart'))
_TABLE_VISIBLE_SHOW_MODES = frozenset(('table', 'show-table', 'chart-table'))
_CHART_FILE_SHOW_MODES = frozenset((
    'show-chart', 'chart', 'chart-table', 'dashboard',
))


def payload_show_mode(payload):
    """Normalized ``show_mode`` / ``showmode`` token, or empty."""
    if not isinstance(payload, dict):
        return ''
    raw = payload.get('show_mode') or payload.get('showmode') or ''
    return str(raw).strip().lower()


def wants_on_screen_table(payload):
    """True when the listing must stay visible next to the card.

    ``__show_table__`` is the explicit flag. ``show-table`` / ``table`` /
    ``chart-table`` also keep the listing. ``show-chart`` alone does not
    (dataset stays for the canvas).
    """
    if not isinstance(payload, dict):
        return False
    if payload.get('__show_table__'):
        return True
    if payload_has_report_outline(payload):
        return True
    return payload_show_mode(payload) in _TABLE_VISIBLE_SHOW_MODES


def clip_include_table(payload):
    """File listing unless this turn is chart-only."""
    if not isinstance(payload, dict):
        return True
    if payload.get('__show_table__'):
        return True
    if payload_show_mode(payload) in _CHART_ONLY_SHOW_MODES:
        return False
    return True


def wants_on_screen_blocks(payload, rows=None):
    """Keep turn HTML when the listing or a chart should paint."""
    return wants_on_screen_table(payload) or clip_include_chart(payload, rows)


def wants_rich_doc(payload):
    """True when the pack asked for a showpiece Word (flag only)."""
    if not isinstance(payload, dict):
        return False
    return bool(payload.get('__rich_doc__'))


def payload_allows_charts(payload):
    """False only on explicit opt-out (``__no_charts__`` / ``charts is False``)."""
    if not isinstance(payload, dict):
        return True
    if payload.get('__no_charts__') is True:
        return False
    if payload.get('charts') is False:
        return False
    return True


def payload_formatted_html(payload):
    """Turn HTML from a RelaxAICode payload, or empty."""
    if not isinstance(payload, dict):
        return ''
    html = payload.get('formatted_text')
    if isinstance(html, str) and html.strip():
        return html
    return ''


def export_hides_on_screen_table(user_message, payload=None):
    """Named download without ``__show_table__``: card only, no server table.

    A ``report_outline`` payload keeps prose (and tables) next to the card.
    """
    if not requested_named_format(user_message):
        return False
    return not wants_on_screen_table(payload)


def requested_named_format(user_message):
    """Parsed format when the utterance names a serializable type, else None."""
    parsed = parse_requested_export(user_message)
    if parsed and parsed.get('format') and not parsed.get('refused'):
        return parsed
    return None


def is_refused_export(user_message):
    """True when every named file type is outside the closed list.

    Mixed utterances (PDF + PowerPoint) stay False so the turn can query
    and serialize the kinds that work.
    """
    parsed = parse_requested_export(user_message)
    if not parsed:
        return False
    formats = parsed.get('formats')
    if isinstance(formats, list) and formats:
        return False
    return bool(parsed.get('refused') and not parsed.get('format'))


def refused_token_labels(tokens):
    """Comma-separated utterance tokens, first spelling wins."""
    seen = set()
    labels = []
    for tok in tokens or []:
        text = ' '.join(str(tok).split())
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        labels.append(text)
    return ', '.join(labels)


def merge_refused_note(body, tokens, note=None):
    """Append the partial-refuse sentence, or return body unchanged."""
    labels = refused_token_labels(tokens)
    if not labels:
        return body or ''
    template = note if note is not None else REFUSED_PARTIAL_REPLY
    text = template % labels
    body = (body or '').rstrip()
    if body:
        return body + '\n\n' + text
    return text


_TITLE_MAX = 120
_CAPTION_MAX = 200
_STEM_MAX = 80
_WIDTH_MIN = 8.0
_WIDTH_MAX = 50.0
_WIDTH_SAMPLE = 80


def _plain_flag(value, max_len):
    if not isinstance(value, str):
        return ''
    text = ' '.join(value.split())
    return text[:max_len]


def export_title(payload=None, fallback='Export'):
    """Interior title from ``__title__`` / stem, else ``fallback``."""
    if isinstance(payload, dict):
        title = _plain_flag(payload.get('__title__'), _TITLE_MAX)
        if title:
            return title
        stem = _plain_flag(payload.get('__filename_stem__'), _STEM_MAX)
        if stem:
            return stem.replace('_', ' ').strip() or (fallback or 'Export')
    return fallback or 'Export'


def has_export_meta(payload=None):
    """True when RelaxAICode named the file (title or stem)."""
    if not isinstance(payload, dict):
        return False
    if _plain_flag(payload.get('__title__'), _TITLE_MAX):
        return True
    return bool(_plain_flag(payload.get('__filename_stem__'), _STEM_MAX))


def sheet_caption(payload=None):
    """One-line Excel caption, or empty (document name is enough)."""
    if not isinstance(payload, dict):
        return ''
    return _plain_flag(payload.get('__sheet_caption__'), _CAPTION_MAX)


def column_labels(payload=None, rows=None):
    """Visible headers: ``__column_labels__`` overlay, else public keys."""
    raw = {}
    if isinstance(payload, dict):
        src = payload.get('__column_labels__')
        if isinstance(src, dict):
            raw = src
    out = {}
    for key in public_columns(rows):
        lab = raw.get(key)
        if lab is None:
            lab = raw.get(str(key))
        if lab not in (None, ''):
            out[key] = str(lab).strip() or str(key)
        else:
            out[key] = str(key)
    return out


def column_widths(rows, labels=None):
    """Excel column widths from header + sampled cell text. No locale list."""
    keys = public_columns(rows)
    labels = labels or {}
    sample = [r for r in (rows or [])[:_WIDTH_SAMPLE] if isinstance(r, dict)]
    widths = []
    for key in keys:
        label = str(labels.get(key) or key)
        longest = len(label)
        numeric = True
        saw_value = False
        for row in sample:
            val = row.get(key)
            if val is None or val == '':
                continue
            saw_value = True
            if isinstance(val, bool) or not isinstance(val, (int, float)):
                numeric = False
            text = _xlsx_plain_text(val) if not isinstance(
                val, (int, float),
            ) or isinstance(val, bool) else ('%s' % val)
            if len(text) > longest:
                longest = len(text)
        width = float(longest + 2)
        if saw_value and numeric:
            width = max(width, 12.0)
        width = min(_WIDTH_MAX, max(_WIDTH_MIN, width))
        widths.append(width)
    return widths


def export_result_chips(result):
    """Download chips from ``apply_requested_exports`` (N or one)."""
    if not isinstance(result, dict) or not result.get('acted'):
        return []
    out = []
    chips = result.get('chips')
    if isinstance(chips, list):
        for chip in chips:
            if isinstance(chip, dict) and (chip.get('url') or chip.get('name')):
                out.append(chip)
    if out:
        return out
    chip = result.get('chip')
    if isinstance(chip, dict) and (chip.get('url') or chip.get('name')):
        return [chip]
    return []


def has_export_chip(result):
    """True when persist returned a download chip the bubble can paint."""
    return bool(export_result_chips(result))


def is_session_file_propose(steps):
    """True when a Safe Plan dumps a file onto orphan ``ir.attachment``.

    Business attachments (``res_model`` = ``res.partner``, invoice, …)
    stay False. Host-testable; no Odoo.
    """
    for step in steps or []:
        if not isinstance(step, dict):
            continue
        op = str(step.get('op') or '').strip().lower()
        if op not in ('create', 'write'):
            continue
        if str(step.get('model') or '').strip() != 'ir.attachment':
            continue
        values = step.get('values')
        if not isinstance(values, dict):
            values = {}
        res_model = values.get('res_model')
        if isinstance(res_model, str):
            res_model = res_model.strip()
        if res_model and res_model != _CHATBOO_SESSION_MODEL:
            continue
        if op == 'create':
            return True
        has_file = any(
            values.get(key) not in (None, '', [])
            for key in _FILE_VALUE_KEYS
        )
        if has_file:
            return True
    return False


def public_columns(rows):
    """Column keys that do not start with ``_``, in first-seen order."""
    keys = []
    seen = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        for key in row:
            if str(key).startswith('_'):
                continue
            if key in seen:
                continue
            seen.add(key)
            keys.append(key)
    return keys


def public_rows(rows):
    keys = public_columns(rows)
    if not keys:
        return []
    out = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        out.append({k: row.get(k) for k in keys})
    return out


def _clip_cell(value):
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (str, int, float)):
        return value
    if isinstance(value, (bytes, bytearray, memoryview)):
        return None
    return str(value)


def clip_data_from_payload(payload=None, rows=None):
    """JSON rows for bubble clip icons (same shape as data-chatboo-dataset)."""
    public = public_rows(tabular_rows(payload, rows))
    if not public:
        return None
    rows_out = []
    for row in public:
        item = {}
        for key, val in row.items():
            cell = _clip_cell(val)
            if cell is not None or val is None:
                item[key] = cell
        if item:
            rows_out.append(item)
    if not rows_out:
        return None
    title = export_title(payload, fallback='')
    out = {
        'rows': rows_out,
        'landscape': bool(pdf_landscape(payload, rows=rows_out)),
        'include_table': clip_include_table(payload),
        'include_chart': clip_include_chart(payload, rows_out),
    }
    if title and title != 'Export':
        out['title'] = title
    return out


def tabular_rows(payload=None, rows=None):
    """First list-of-dicts from an explicit list or a RelaxAICode payload."""
    if isinstance(rows, list) and rows and isinstance(rows[0], dict):
        return list(rows)
    if isinstance(payload, list) and payload and isinstance(payload[0], dict):
        return list(payload)
    if not isinstance(payload, dict):
        return None
    data = payload.get('data')
    if isinstance(data, list) and data and isinstance(data[0], dict):
        return list(data)
    items = payload.get('items')
    if isinstance(items, list) and items and isinstance(items[0], dict):
        return list(items)
    for envelope in ('groups', 'sections', 'tables'):
        blocks = payload.get(envelope)
        if not isinstance(blocks, list):
            continue
        for block in blocks:
            if not isinstance(block, dict):
                continue
            for key in ('data', 'items', 'rows'):
                nested = block.get(key)
                if isinstance(nested, list) and nested and isinstance(nested[0], dict):
                    return list(nested)
    for key, val in payload.items():
        if str(key).startswith('_'):
            continue
        if isinstance(val, list) and val and isinstance(val[0], dict):
            return list(val)
    return None


def columns_config(rows, labels=None):
    """``format_data_as_excel`` / ``format_data_as_pdf`` tuples."""
    labels = labels or {}
    cfg = []
    for key in public_columns(rows):
        label = labels.get(key)
        cfg.append((key, str(label if label not in (None, '') else key), 'text', 'left'))
    return cfg


_PDF_CHART_MIN_ROWS = 3
_PDF_CHART_MAX_CATS = 24
_PDF_CHART_MAX_SERIES = 4
_PDF_CHART_SCALE_RATIO = 8.0


def _is_chart_id_key(key):
    name = str(key or '')
    return name == 'id' or name.endswith('_id')


def coerce_chart_number(value):
    """Parse a cell into a float. Separators only; no locale word list."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value != value or value in (float('inf'), float('-inf')):
            return None
        return float(value)
    text = str(value or '').strip()
    if not text:
        return None
    cleaned = []
    for ch in text:
        if ch.isdigit() or ch in '.,-':
            cleaned.append(ch)
    text = ''.join(cleaned)
    if text in ('', '-', '--', '.', ',', '-.', '-,'):
        return None
    if text.count(',') == 1 and text.count('.') > 1:
        text = text.replace('.', '').replace(',', '.')
    elif text.count(',') == 1 and text.count('.') == 0:
        text = text.replace(',', '.')
    elif text.count(',') == 1 and text.count('.') == 1:
        if text.rfind(',') > text.rfind('.'):
            text = text.replace('.', '').replace(',', '.')
        else:
            text = text.replace(',', '')
    try:
        return float(text)
    except ValueError:
        return None


def _column_is_numeric(rows, key):
    hits = 0
    seen = 0
    for row in rows:
        val = row.get(key) if isinstance(row, dict) else None
        if val in (None, ''):
            continue
        seen += 1
        if coerce_chart_number(val) is not None:
            hits += 1
    return seen >= 2 and hits >= max(2, int(seen * 0.7))


def payload_has_chart_series(payload=None, rows=None):
    """True when public rows have at least two columns and one numeric series."""
    public = public_rows(tabular_rows(payload, rows) or [])
    if not public:
        return False
    keys = public_columns(public)
    if len(keys) < 2:
        return False
    return any(_column_is_numeric(public, key) for key in keys)


def clip_include_chart(payload, rows=None):
    """File/bubble chart when the turn asked for a graphic view and series exist."""
    if not payload_allows_charts(payload):
        return False
    mode = payload_show_mode(payload)
    if mode not in _CHART_FILE_SHOW_MODES:
        return False
    return payload_has_chart_series(payload, rows)


def pdf_chart_series(rows, labels=None):
    """Category + 1–4 numeric series, or None. Shape only (no domain lexicon)."""
    public = public_rows(rows)
    if not public or len(public) < _PDF_CHART_MIN_ROWS:
        return None
    if len(public) > _PDF_CHART_MAX_CATS:
        return None
    keys = public_columns(public)
    if len(keys) < 2:
        return None
    cat_key = None
    for key in keys:
        if _is_chart_id_key(key):
            continue
        cat_key = key
        break
    if not cat_key:
        return None
    numeric = [
        key for key in keys
        if key != cat_key
        and not _is_chart_id_key(key)
        and _column_is_numeric(public, key)
    ][:_PDF_CHART_MAX_SERIES]
    if not numeric:
        return None
    categories = []
    raw_series = {key: [] for key in numeric}
    for row in public:
        cat = row.get(cat_key)
        if cat is None or str(cat).strip() == '':
            return None
        categories.append(str(cat).strip())
        for key in numeric:
            raw_series[key].append(coerce_chart_number(row.get(key)))
    labels = labels or {}
    series = []
    for key in numeric:
        values = raw_series[key]
        if all(val is None for val in values):
            continue
        series.append({
            'name': str(labels.get(key) or key),
            'values': [0.0 if val is None else val for val in values],
        })
    if not series:
        return None
    maxes = [max(abs(val) for val in item['values']) for item in series]
    peak = max(maxes) if maxes else 0.0
    if peak > 0:
        kept = []
        dropped_scale = False
        for item, magnitude in zip(series, maxes):
            if magnitude <= 0:
                continue
            ratio = peak / max(magnitude, 1e-9)
            if ratio <= _PDF_CHART_SCALE_RATIO:
                kept.append(item)
            else:
                dropped_scale = True
        if dropped_scale:
            return None
        series = kept
        if not series:
            return None
    return {
        'categories': categories,
        'series': series,
        'kind': 'line' if len(categories) >= 8 else 'bar',
    }


def serialize_rows(rows, kind, title='Export', landscape=True, labels=None,
                   caption=None, env=None, html=None, charts=True):
    """Return bytes for ``kind`` (txt/md/html/xlsx/pdf/doc), or None."""
    public = public_rows(rows)
    kind = (kind or '').lower()
    labels = labels or {}
    if kind == 'doc':
        return _as_word_html(
            public, title, labels=labels, html=html, env=env,
        )
    if not public:
        return None
    if kind == 'txt':
        return _as_tsv(public).encode('utf-8')
    if kind == 'md':
        return _as_markdown(public, title=title, labels=labels).encode('utf-8')
    if kind == 'html':
        return _as_html(public, title, labels=labels).encode('utf-8')
    if kind == 'xlsx':
        return _as_xlsx(public, title, labels=labels, caption=caption)
    if kind == 'pdf':
        return _as_pdf(
            public, title, landscape=landscape, labels=labels, env=env,
            charts=charts,
        )
    return None


def _slug_stem(text):
    try:
        svg = _load_sibling('svg_download.py', '_ae_svg_download')
        slug = svg._slug(text or '')
        if slug:
            return slug[:_STEM_MAX]
    except Exception:
        _logger.debug('artifact_export: slug fallback', exc_info=True)
    slug = re.sub(r'[^A-Za-z0-9]+', '_', str(text or '')).strip('_').lower()
    return slug[:_STEM_MAX]


def export_filename(prompt, ext, fallback='export', payload=None, title=None):
    """Prefer payload stem/title; else slug the utterance."""
    ext = (ext or 'bin').lstrip('.').lower() or 'bin'
    stem = ''
    if isinstance(payload, dict):
        stem = _slug_stem(_plain_flag(payload.get('__filename_stem__'), _STEM_MAX))
        if not stem:
            stem = _slug_stem(_plain_flag(payload.get('__title__'), _TITLE_MAX))
    if not stem and title:
        stem = _slug_stem(title)
    if stem:
        return '%s.%s' % (stem, ext)
    try:
        svg = _load_sibling('svg_download.py', '_ae_svg_download')
        name = svg.utterance_filename(
            prompt=prompt or '',
            ext=ext,
            fallback=fallback,
        )
        if name:
            return name
    except Exception:
        _logger.debug('artifact_export: utterance_filename fallback', exc_info=True)
    return '%s.%s' % ((fallback or 'export'), ext)


def file_banner_tone(chip):
    """CSS suffix: pdf / excel / text / code / other."""
    chip = chip or {}
    mt = (chip.get('mimetype') or '').split(';', 1)[0].strip().lower()
    name = (chip.get('name') or '').lower()
    ext = name.rsplit('.', 1)[-1] if '.' in name else ''
    if mt == 'application/pdf' or ext == 'pdf':
        return 'pdf'
    if (
        mt == 'application/msword'
        or 'wordprocessing' in mt
        or ext in ('doc', 'docx')
    ):
        return 'word'
    if (
        'spreadsheet' in mt
        or 'excel' in mt
        or mt == 'text/csv'
        or ext in ('xls', 'xlsx', 'ods', 'csv')
    ):
        return 'excel'
    if (
        'json' in mt
        or 'xml' in mt
        or 'javascript' in mt
        or 'html' in mt
        or ext in ('json', 'xml', 'js', 'html', 'css', 'py')
    ):
        return 'code'
    if mt.startswith('text/') or ext in ('txt', 'md', 'log', 'markdown'):
        return 'text'
    return 'other'


def file_icon_class(chip):
    tone = file_banner_tone(chip)
    mapping = {
        'pdf': 'fa fa-file-pdf-o',
        'excel': 'fa fa-file-excel-o',
        'word': 'fa fa-file-word-o',
        'code': 'fa fa-file-code-o',
        'text': 'fa fa-file-text-o',
    }
    return mapping.get(tone, 'fa fa-file-o')


def format_file_size(n):
    if n is None or n == '':
        return ''
    try:
        n = int(n)
    except (TypeError, ValueError):
        return ''
    if n < 1024:
        return '%d B' % n
    if n < 1024 * 1024:
        return '%.1f KB' % (n / 1024.0)
    return '%.1f MB' % (n / (1024.0 * 1024.0))


def file_banner_html(chip, cta=DOWNLOAD_CTA):
    """Markup matching ``.o_chatboo_file_banner_*`` (tests + server fallback)."""
    chip = chip or {}
    href = html_lib.escape(str(chip.get('url') or '#'), quote=True)
    name = html_lib.escape(str(chip.get('name') or 'download'))
    tone = file_banner_tone(chip)
    icon = file_icon_class(chip)
    size = format_file_size(chip.get('size'))
    size_html = (
        '<span class="o_chatboo_file_banner_size">%s</span>' % html_lib.escape(size)
        if size else ''
    )
    cta_txt = html_lib.escape(cta or DOWNLOAD_CTA)
    return (
        '<a class="o_chatboo_file_banner_card o_chatboo_file_banner_%s" '
        'href="%s" target="_blank" rel="noopener" download="%s">'
        '<span class="o_chatboo_file_banner_preview" aria-hidden="true">'
        '<i class="%s"></i></span>'
        '<span class="o_chatboo_file_banner_meta">'
        '<span class="o_chatboo_file_banner_title">%s</span>'
        '%s'
        '<span class="o_chatboo_file_banner_cta">%s</span>'
        '</span></a>'
    ) % (tone, href, name, icon, name, size_html, cta_txt)


def _parsed_format_specs(parsed):
    """Unique format dicts in appearance order (compat: first hit)."""
    if not parsed:
        return []
    specs = parsed.get('formats')
    if isinstance(specs, list) and specs:
        return [s for s in specs if isinstance(s, dict) and s.get('kind')]
    if parsed.get('format') and not parsed.get('refused'):
        return [parsed]
    return []


def apply_requested_exports(
    user_message,
    rows=None,
    payload=None,
    persist=None,
    prompt=None,
    env=None,
    client_fulfill=False,
):
    """Parse + serialize every unique kind + optional persist.

    One chip per serializable kind; refused tokens stay on the result
    without blocking the rest. Never raises for bad utterances.
    ``persist(raw, filename, mimetype) -> chip|None``. When omitted
    (host tests), a synthetic chip is returned per kind.

    ``client_fulfill``: Word / PDF / HTML become pending chips. Chatboo
    assembles the real file (table + painted chart) and replaces them.
    Excel and Markdown still serialize here.
    """
    empty = {
        'parsed': None,
        'acted': False,
        'chip': None,
        'chips': [],
        'raw': None,
        'filename': None,
        'refused_tokens': [],
    }
    parsed = parse_requested_export(user_message)
    empty['parsed'] = parsed
    if parsed:
        empty['refused_tokens'] = list(parsed.get('refused_tokens') or [])
    specs = _parsed_format_specs(parsed)
    if not specs:
        return empty
    row_list = tabular_rows(payload, rows)
    public = public_rows(row_list)
    html_body = payload_formatted_html(payload)
    rich = wants_rich_doc(payload)
    has_doc = any(s.get('kind') == 'doc' for s in specs)
    if not public:
        if not (rich and html_body and has_doc):
            return empty
        specs = [s for s in specs if s.get('kind') == 'doc']
    title = export_title(payload, fallback='Export')
    labels = column_labels(payload, public)
    caption = sheet_caption(payload)
    landscape = pdf_landscape(
        payload,
        col_count=len(public_columns(public)) if public else None,
        rows=public,
    )
    chips = []
    last_raw = None
    last_filename = None
    for spec in specs:
        filename = export_filename(
            prompt or user_message, spec['ext'],
            payload=payload, title=title,
        )
        if client_fulfill and spec.get('kind') in CLIENT_ASSEMBLED_KINDS:
            chip = {
                'name': filename,
                'mimetype': spec['mimetype'],
                'size': 0,
                'pending': True,
                'fulfill': spec['kind'],
                'source': 'download',
            }
            chips.append(chip)
            last_filename = filename
            continue
        extra_html = html_body if (spec.get('kind') == 'doc' and rich) else None
        raw = serialize_rows(
            public, spec['kind'], title=title,
            landscape=landscape, labels=labels, caption=caption, env=env,
            html=extra_html,
            charts=payload_allows_charts(payload),
        )
        if not raw:
            continue
        chip = None
        if persist is not None:
            chip = persist(raw, filename, spec['mimetype'])
            if not chip:
                continue
        else:
            chip = {
                'name': filename,
                'url': '#',
                'mimetype': spec['mimetype'],
                'size': len(raw),
            }
        chips.append(chip)
        last_raw = raw
        last_filename = filename
    if not chips:
        return empty
    return {
        'parsed': parsed,
        'acted': True,
        'chip': chips[0],
        'chips': chips,
        'raw': last_raw,
        'filename': last_filename,
        'refused_tokens': list(parsed.get('refused_tokens') or []),
    }


def apply_requested_export(
    user_message,
    rows=None,
    payload=None,
    persist=None,
    prompt=None,
    env=None,
    client_fulfill=False,
):
    """Parse + serialize + optional persist. Never raises for bad utterances.

    Table vs card is the engine (``export_hides_on_screen_table``). PPT
    does not serialize. One card per unique kind when several are named.
    ``persist(raw, filename, mimetype) -> chip|None``. When omitted
    (host tests), a synthetic chip is returned. When provided and every
    persist returns None, ``acted`` stays False.
    """
    return apply_requested_exports(
        user_message,
        rows=rows,
        payload=payload,
        persist=persist,
        prompt=prompt,
        env=env,
        client_fulfill=client_fulfill,
    )


def _as_tsv(rows):
    keys = public_columns(rows)
    lines = ['\t'.join(_cell(k) for k in keys)]
    for row in rows:
        lines.append('\t'.join(_cell(row.get(k)) for k in keys))
    return '\n'.join(lines) + '\n'


def _as_markdown(rows, title=None, labels=None):
    keys = public_columns(rows)
    labels = labels or {}
    heads = [_cell(labels.get(k) or k) for k in keys]
    header = '| ' + ' | '.join(heads) + ' |'
    sep = '| ' + ' | '.join('---' for _k in keys) + ' |'
    body = [
        '| ' + ' | '.join(_cell(row.get(k)) for k in keys) + ' |'
        for row in rows
    ]
    parts = []
    if title:
        parts.append('# %s' % _cell(title))
        parts.append('')
    parts.extend([header, sep] + body)
    return '\n'.join(parts) + '\n'


def _as_html(rows, title, labels=None):
    keys = public_columns(rows)
    labels = labels or {}
    title_h = html_lib.escape(str(title or 'Export'))
    thead = ''.join(
        '<th>%s</th>' % html_lib.escape(_cell(labels.get(k) or k))
        for k in keys
    )
    body_rows = []
    for row in rows:
        tds = ''.join(
            '<td>%s</td>' % html_lib.escape(_cell(row.get(k))) for k in keys
        )
        body_rows.append('<tr>%s</tr>' % tds)
    return (
        '<!DOCTYPE html><html><head><meta charset="utf-8"/>'
        '<title>%s</title></head><body><h1>%s</h1>'
        '<table><thead><tr>%s</tr></thead><tbody>%s</tbody></table>'
        '</body></html>'
    ) % (title_h, title_h, thead, ''.join(body_rows))


_WORD_KILL_CLASS_RE = re.compile(
    r'class="[^"]*(?:o_chatboo_chart_toolbar|o_chatboo_export_bar|'
    r'o_chatboo_copy_btn|o_chatboo_noexport|o_chatboo_ts|'
    r'o_chatboo_chart_host|o_chatboo_echarts_surface|echarts-tooltip)[^"]*"',
    re.I,
)
_WORD_TAG_RE = re.compile(r'<(/?)([a-zA-Z][\w:-]*)([^>]*)>', re.S)
_WORD_VOID = frozenset((
    'area', 'base', 'br', 'col', 'embed', 'hr', 'img', 'input',
    'link', 'meta', 'param', 'source', 'track', 'wbr',
))
_WORD_SCRIPT_RE = re.compile(r'<script[\s\S]*?</script>', re.I)
_WORD_STYLE = (
    'body{font-family:Helvetica,Arial,sans-serif;color:#1a1a1a;font-size:11pt;}'
    'h1{font-size:16pt;color:#1a1a1a;margin:0 0 12pt;}'
    'table{border-collapse:collapse;width:100%;}'
    'th{background:#eceff1;color:#1a1a1a;font-weight:bold;'
    'padding:6px 8px;border:0.5pt solid #c8c8c8;text-align:left;}'
    'td{padding:5px 8px;border:0.5pt solid #c8c8c8;color:#282828;'
    'vertical-align:middle;}'
    'tr:nth-child(even) td{background:#f8f9fa;}'
    'td img{max-width:72px;max-height:72px;}'
    'p.o_chatboo_word_figure img{max-width:100%;width:100%;height:auto;}'
)


def _word_matching_close(text, start, name):
    """Index after the close tag that matches ``name`` at ``start``."""
    depth = 1
    name_l = name.lower()
    for match in _WORD_TAG_RE.finditer(text, start):
        slash, other, attrs = match.group(1), match.group(2), match.group(3)
        if other.lower() != name_l:
            continue
        if slash:
            depth -= 1
            if depth == 0:
                return match.end()
            continue
        if (attrs or '').rstrip().endswith('/') or other.lower() in _WORD_VOID:
            continue
        depth += 1
    return len(text)


def _word_strip_kill_nodes(html):
    """Drop chart chrome / tooltips with matching close tags (not first ``</``)."""
    text = str(html or '')
    while True:
        found = None
        for match in _WORD_TAG_RE.finditer(text):
            slash, name, attrs = match.group(1), match.group(2), match.group(3)
            if slash or not _WORD_KILL_CLASS_RE.search(attrs or ''):
                continue
            if (attrs or '').rstrip().endswith('/') or name.lower() in _WORD_VOID:
                found = (match.start(), match.end())
                break
            end = _word_matching_close(text, match.end(), name)
            found = (match.start(), end)
            break
        if not found:
            return text
        text = text[:found[0]] + text[found[1]:]


def _word_sanitize_html(html):
    text = str(html or '')
    text = _WORD_SCRIPT_RE.sub('', text)
    text = _word_strip_kill_nodes(text)
    return text.strip()


def _word_wrap(title, body):
    title_h = html_lib.escape(str(title or 'Export'))
    return (
        '<html xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:w="urn:schemas-microsoft-com:office:word">'
        '<head><meta charset="utf-8"/><title>%s</title>'
        '<style>%s</style></head><body>%s</body></html>'
    ) % (title_h, _WORD_STYLE, body)


def _word_cell_html(value, key, env=None):
    fmt = _load_formatters()
    if fmt is not None:
        reader = getattr(fmt, '_pdf_image_bytes', None)
        if callable(reader):
            raw = reader(value, key=key, env=env)
            if raw:
                import base64
                b64 = base64.b64encode(raw).decode('ascii')
                return '<img src="data:image/png;base64,%s" alt=""/>' % b64
        skip = getattr(fmt, '_pdf_is_image_ref', None)
        if callable(skip) and skip(value):
            return ''
    return html_lib.escape(_cell(value))


def _word_table_html(rows, title, labels=None, env=None):
    keys = public_columns(rows)
    labels = labels or {}
    title_h = html_lib.escape(str(title or 'Export'))
    thead = ''.join(
        '<th>%s</th>' % html_lib.escape(_cell(labels.get(k) or k))
        for k in keys
    )
    body_rows = []
    for row in rows:
        tds = ''.join(
            '<td>%s</td>' % _word_cell_html(row.get(k), k, env=env)
            for k in keys
        )
        body_rows.append('<tr>%s</tr>' % tds)
    return (
        '<h1>%s</h1><table><thead><tr>%s</tr></thead><tbody>%s</tbody></table>'
    ) % (title_h, thead, ''.join(body_rows))


def _as_word_html(rows, title, labels=None, html=None, env=None):
    """Office Word HTML (.doc). Rich body or Chatboo table. No new deps."""
    body = ''
    if html and str(html).strip():
        body = _word_sanitize_html(html)
        if body and '<h1' not in body.lower() and title:
            body = '<h1>%s</h1>%s' % (
                html_lib.escape(str(title)), body,
            )
    elif rows:
        body = _word_table_html(rows, title, labels=labels, env=env)
    if not body:
        return None
    return _word_wrap(title, body).encode('utf-8')


_MAX_XLSX_SHEETS = 32
_MAX_XLSX_ROWS = 20000
_MAX_XLSX_COLS = 256


def _as_xlsx(rows, title, labels=None, caption=None):
    """Prefer styled openpyxl; always fall back to stdlib OOXML."""
    return serialize_xlsx_sheets([{
        'title': title,
        'rows': rows,
        'labels': labels,
        'caption': caption,
    }])


def serialize_xlsx_sheets(sheets):
    """One workbook from ``[{title, rows, labels, caption}]``."""
    specs = _normalize_xlsx_sheets(sheets)
    if not specs:
        return None
    fmt = _load_formatters()
    if fmt is not None and hasattr(fmt, 'format_excel_workbook'):
        try:
            raw = fmt.format_excel_workbook([
                {
                    'data': spec['rows'],
                    'columns_config': spec['cfg'],
                    'title': spec['title'],
                    'include_row_index': False,
                    'caption': spec['caption'] or None,
                    'widths': spec['widths'],
                }
                for spec in specs
            ])
            coerced = _coerce_xlsx_bytes(raw)
            if coerced:
                return coerced
        except Exception:
            _logger.debug(
                'artifact_export: format_excel_workbook failed', exc_info=True,
            )
    if fmt is not None and len(specs) == 1:
        spec = specs[0]
        try:
            raw = fmt.format_data_as_excel(
                spec['rows'], spec['cfg'], title=spec['title'],
                include_row_index=False, caption=spec['caption'] or None,
                widths=spec['widths'],
            )
            coerced = _coerce_xlsx_bytes(raw)
            if coerced:
                return coerced
        except Exception:
            _logger.debug(
                'artifact_export: format_data_as_excel failed', exc_info=True,
            )
    try:
        raw = _as_xlsx_openpyxl_book(specs)
        coerced = _coerce_xlsx_bytes(raw)
        if coerced:
            return coerced
    except ImportError:
        _logger.info(
            'artifact_export: openpyxl not installed; using stdlib xlsx',
        )
    except Exception:
        _logger.debug('artifact_export: openpyxl write failed', exc_info=True)
    try:
        return _as_xlsx_ooxml_book(specs)
    except Exception:
        _logger.warning('artifact_export: stdlib xlsx failed', exc_info=True)
        return None


def sheets_from_aoa_sections(sections):
    """Icon payload ``[{title, aoa}]`` → serialize_xlsx_sheets specs."""
    out = []
    used = set()
    for spec in (sections or [])[:_MAX_XLSX_SHEETS]:
        if not isinstance(spec, dict):
            continue
        rows_in = spec.get('rows')
        if (
            isinstance(rows_in, list)
            and rows_in
            and isinstance(rows_in[0], dict)
        ):
            rows = public_rows(rows_in[:_MAX_XLSX_ROWS])
            labels = spec.get('labels') or {
                k: str(k) for k in public_columns(rows)
            }
            if not (rows or labels):
                continue
            out.append({
                'title': _unique_sheet_title(spec.get('title'), used),
                'rows': rows,
                'labels': labels,
                'caption': (spec.get('caption') or '').strip(),
            })
            continue
        aoa = spec.get('aoa')
        parsed = _sheet_from_aoa(aoa, spec.get('title'), used)
        if parsed:
            parsed['caption'] = (spec.get('caption') or '').strip()
            out.append(parsed)
    return out


def icon_xlsx_payload(sections, filename=None):
    """Base64 download dict for the Chatboo Excel icon, or None."""
    raw = serialize_xlsx_sheets(sheets_from_aoa_sections(sections))
    if not raw:
        return None
    return {
        'filename': _safe_xlsx_filename(filename),
        'mimetype': _XLSX_MIME,
        'datas': base64.b64encode(raw).decode('ascii'),
    }


def _sheet_from_aoa(aoa, title, used):
    if not isinstance(aoa, list) or not aoa:
        return None
    header = [_aoa_cell(c) for c in (aoa[0] or [])][:_MAX_XLSX_COLS]
    keys = []
    seen = {}
    labels = {}
    for idx, raw_h in enumerate(header):
        label = str(raw_h).strip() if raw_h not in (None, '') else ''
        key = label or ('Col%d' % (idx + 1))
        n = seen.get(key, 0)
        seen[key] = n + 1
        if n:
            key = '%s_%d' % (key, n + 1)
        keys.append(key)
        labels[key] = label or key
    if not keys:
        return None
    rows = []
    for raw in aoa[1:_MAX_XLSX_ROWS + 1]:
        if not isinstance(raw, (list, tuple)):
            continue
        row = {}
        for idx, key in enumerate(keys):
            row[key] = _aoa_cell(raw[idx]) if idx < len(raw) else ''
        rows.append(row)
    return {
        'title': _unique_sheet_title(title, used),
        'rows': rows,
        'labels': labels,
    }


def _aoa_cell(value):
    if isinstance(value, dict):
        if 'content' in value:
            return value.get('content')
        return ''
    return value


def _safe_xlsx_filename(filename):
    text = str(filename or '').strip() or 'export'
    text = text.replace('\\', '/').split('/')[-1]
    if '.' in text:
        text = text.rsplit('.', 1)[0]
    stem = _slug_stem(text) or 'export'
    return '%s.xlsx' % stem


def _normalize_xlsx_sheets(sheets):
    specs = []
    used = set()
    for spec in (sheets or [])[:_MAX_XLSX_SHEETS]:
        if not isinstance(spec, dict):
            continue
        labels_in = spec.get('labels') or {}
        rows = public_rows((spec.get('rows') or [])[:_MAX_XLSX_ROWS])
        keys = public_columns(rows)
        if not keys:
            keys = [
                k for k in labels_in
                if not str(k).startswith('_')
            ][:_MAX_XLSX_COLS]
        if not keys:
            continue
        labels = {
            k: str(labels_in.get(k) or k) for k in keys
        }
        widths = spec.get('widths') or column_widths(
            rows or [{k: labels[k] for k in keys}], labels,
        )
        typed = [
            {k: _excel_typed(row.get(k)) for k in keys}
            for row in rows
        ]
        specs.append({
            'title': _unique_sheet_title(spec.get('title'), used),
            'rows': typed,
            'labels': labels,
            'caption': (spec.get('caption') or '').strip(),
            'widths': widths,
            'keys': keys,
            'cfg': columns_config(typed or [{k: '' for k in keys}], labels),
        })
    return specs


def _as_xlsx_openpyxl_book(specs):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    first = True
    for spec in specs:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = spec['title']
        keys = spec['keys']
        labels = spec['labels']
        caption = spec['caption']
        header_row = 2 if caption else 1
        if caption:
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=max(1, len(keys)),
            )
            ws.cell(row=1, column=1, value=_xlsx_plain_text(caption))
        for col, key in enumerate(keys, 1):
            ws.cell(
                row=header_row, column=col,
                value=str(labels.get(key) or key),
            )
        for ridx, row in enumerate(spec['rows'], header_row + 1):
            for col, key in enumerate(keys, 1):
                ws.cell(
                    row=ridx, column=col,
                    value=_excel_value(row.get(key)),
                )
        for col, width in enumerate(spec['widths'], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = 'A%d' % (header_row + 1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _as_pdf(rows, title, landscape=True, labels=None, env=None, charts=True):
    fmt = _load_formatters()
    cfg = columns_config(rows, labels)
    chart_series = pdf_chart_series(rows, labels) if charts else None
    if fmt is not None:
        try:
            return fmt.format_data_as_pdf(
                rows, cfg, title=title or 'Export', landscape=landscape,
                env=env, chart_series=chart_series,
            )
        except TypeError:
            try:
                return fmt.format_data_as_pdf(
                    rows, cfg, title=title or 'Export', landscape=landscape,
                    env=env,
                )
            except TypeError:
                try:
                    return fmt.format_data_as_pdf(
                        rows, cfg, title=title or 'Export', landscape=landscape,
                    )
                except TypeError:
                    try:
                        return fmt.format_data_as_pdf(
                            rows, cfg, title=title or 'Export',
                        )
                    except Exception:
                        _logger.debug(
                            'artifact_export: format_data_as_pdf failed',
                            exc_info=True,
                        )
                except Exception:
                    _logger.debug(
                        'artifact_export: format_data_as_pdf failed',
                        exc_info=True,
                    )
            except Exception:
                _logger.debug('artifact_export: format_data_as_pdf failed', exc_info=True)
        except Exception:
            _logger.debug('artifact_export: format_data_as_pdf failed', exc_info=True)
    return None


def _excel_number(value):
    """Native int/float only. A digit string stays text (ids, codes)."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value or value in (float('inf'), float('-inf')):
            return None
        return value
    return None


def _excel_typed(value):
    number = _excel_number(value)
    return value if number is None else number


def _excel_value(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return value
    number = _excel_number(value)
    if number is not None:
        return number
    return _xlsx_plain_text(value)


def _coerce_xlsx_bytes(raw):
    """Accept bytes / bytearray / BytesIO; require a ZIP signature."""
    if hasattr(raw, 'getvalue') and not isinstance(raw, (bytes, bytearray)):
        try:
            raw.seek(0)
        except Exception:
            pass
        try:
            raw = raw.getvalue()
        except Exception:
            return None
    if isinstance(raw, memoryview):
        raw = raw.tobytes()
    if isinstance(raw, bytearray):
        raw = bytes(raw)
    if isinstance(raw, bytes) and raw[:2] == b'PK':
        return raw
    return None


_XLSX_SHEET_BAD = re.compile(r'[:\\/\?\*\[\]]')
_XLSX_NS_MAIN = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_XLSX_NS_PKG_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
_XLSX_NS_OD_REL = (
    'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
)
_XLSX_NS_TYPES = 'http://schemas.openxmlformats.org/package/2006/content-types'


def _xlsx_sheet_title(title):
    name = _XLSX_SHEET_BAD.sub(' ', str(title or 'Export')).strip() or 'Export'
    name = name.strip("'") or 'Export'
    return name[:31]


def _unique_sheet_title(title, used):
    base = _xlsx_sheet_title(title)
    name = base
    n = 1
    while name in used:
        suffix = '_%d' % n
        name = (base[:max(1, 31 - len(suffix))] + suffix)[:31]
        n += 1
    used.add(name)
    return name


def _xlsx_frozen_header_xml(header_row):
    """Freeze every row above the first data row (caption + column names)."""
    split = int(header_row)
    return (
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="%d" topLeftCell="A%d" '
        'activePane="bottomLeft" state="frozen"/>'
        '</sheetView></sheetViews>'
    ) % (split, split + 1)


def _xlsx_col_letter(index):
    """1-based column index → A, B, …, Z, AA."""
    n = int(index)
    letters = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(65 + rem))
    return ''.join(reversed(letters)) or 'A'


def _xlsx_plain_text(value):
    if value is None:
        return ''
    text = str(value)
    return ''.join(
        ch for ch in text if ord(ch) >= 32 or ch in '\t\n\r'
    )[:32767]


def _xlsx_xml_text(value):
    return html_lib.escape(_xlsx_plain_text(value), quote=True)


def _xlsx_cell_xml(col, row, value):
    ref = '%s%s' % (_xlsx_col_letter(col), row)
    if isinstance(value, bool):
        return '<c r="%s" t="b"><v>%d</v></c>' % (ref, 1 if value else 0)
    if isinstance(value, int) and not isinstance(value, bool):
        return '<c r="%s"><v>%d</v></c>' % (ref, value)
    if isinstance(value, float):
        if value != value or value in (float('inf'), float('-inf')):
            value = ''
        else:
            return '<c r="%s"><v>%s</v></c>' % (ref, repr(value))
    text = _xlsx_xml_text(value)
    return (
        '<c r="%s" t="inlineStr"><is><t xml:space="preserve">%s</t></is></c>'
        % (ref, text)
    )


def _as_xlsx_ooxml(rows, title='Export', labels=None, caption=None, widths=None):
    """Minimal xlsx via zipfile (no openpyxl). Host- and worker-safe."""
    specs = _normalize_xlsx_sheets([{
        'title': title,
        'rows': rows,
        'labels': labels,
        'caption': caption,
        'widths': widths,
    }])
    if not specs:
        return None
    return _as_xlsx_ooxml_book(specs)


def _xlsx_sheet_xml(spec):
    keys = spec['keys']
    labels = spec['labels']
    caption = spec['caption']
    widths = spec['widths']
    rows = spec['rows']
    header_row = 2 if caption else 1
    col_bits = []
    for idx, width in enumerate(widths, 1):
        col_bits.append(
            '<col min="%d" max="%d" width="%.2f" customWidth="1"/>'
            % (idx, idx, float(width))
        )
    cols_xml = '<cols>%s</cols>' % ''.join(col_bits) if col_bits else ''
    freeze_xml = _xlsx_frozen_header_xml(header_row)
    body = []
    merge_xml = ''
    if caption:
        last = _xlsx_col_letter(len(keys))
        body.append(
            '<row r="1">%s</row>' % _xlsx_cell_xml(1, 1, caption)
        )
        merge_xml = (
            '<mergeCells count="1"><mergeCell ref="A1:%s1"/></mergeCells>'
            % last
        )
    header = ''.join(
        _xlsx_cell_xml(col, header_row, str(labels.get(key) or key))
        for col, key in enumerate(keys, 1)
    )
    body.append('<row r="%d">%s</row>' % (header_row, header))
    for ridx, row in enumerate(rows or [], header_row + 1):
        if not isinstance(row, dict):
            continue
        cells = ''.join(
            _xlsx_cell_xml(col, ridx, row.get(key))
            for col, key in enumerate(keys, 1)
        )
        body.append('<row r="%d">%s</row>' % (ridx, cells))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="%s">%s%s<sheetData>%s</sheetData>%s</worksheet>'
    ) % (_XLSX_NS_MAIN, freeze_xml, cols_xml, ''.join(body), merge_xml)


def _as_xlsx_ooxml_book(specs):
    if not specs:
        return None
    sheet_tags = []
    rel_tags = []
    override_tags = []
    files = []
    for idx, spec in enumerate(specs, 1):
        name = html_lib.escape(spec['title'], quote=True)
        sheet_tags.append(
            '<sheet name="%s" sheetId="%d" r:id="rId%d"/>' % (name, idx, idx)
        )
        rel_tags.append(
            '<Relationship Id="rId%d" Type="%s/worksheet" '
            'Target="worksheets/sheet%d.xml"/>'
            % (idx, _XLSX_NS_OD_REL, idx)
        )
        override_tags.append(
            '<Override PartName="/xl/worksheets/sheet%d.xml" ContentType="'
            'application/vnd.openxmlformats-officedocument.spreadsheetml.'
            'worksheet+xml"/>' % idx
        )
        files.append(
            ('xl/worksheets/sheet%d.xml' % idx, _xlsx_sheet_xml(spec))
        )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="%s" xmlns:r="%s"><sheets>%s</sheets></workbook>'
    ) % (_XLSX_NS_MAIN, _XLSX_NS_OD_REL, ''.join(sheet_tags))
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="%s">%s</Relationships>'
    ) % (_XLSX_NS_PKG_REL, ''.join(rel_tags))
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="%s">'
        '<Relationship Id="rId1" Type="%s/officeDocument" '
        'Target="xl/workbook.xml"/>'
        '</Relationships>'
    ) % (_XLSX_NS_PKG_REL, _XLSX_NS_OD_REL)
    ctypes = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="%s">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-'
        'package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.'
        'openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '%s</Types>'
    ) % (_XLSX_NS_TYPES, ''.join(override_tags))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        def _put(name, data):
            info = zipfile.ZipInfo(filename=name, date_time=(2020, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            zf.writestr(info, data)
        _put('[Content_Types].xml', ctypes)
        _put('_rels/.rels', root_rels)
        _put('xl/workbook.xml', workbook_xml)
        _put('xl/_rels/workbook.xml.rels', workbook_rels)
        for name, data in files:
            _put(name, data)
    return buf.getvalue()


def _cell(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    text = str(value).replace('\t', ' ').replace('\n', ' ')
    return text


def _load_formatters():
    try:
        return _load_path(
            Path(__file__).resolve().parents[1] / 'controllers' / 'formatters.py',
            '_ae_formatters',
        )
    except Exception:
        return None


def _load_sibling(filename, mod_name):
    return _load_path(Path(__file__).resolve().parent / filename, mod_name)


def _load_path(path, mod_name):
    spec = importlib.util.spec_from_file_location(mod_name, path)
    if spec is None or spec.loader is None:
        raise ImportError('cannot load %s' % path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod
