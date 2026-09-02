# -*- coding: utf-8 -*-
# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025-2026 Patanegra Soft <https://patanegra.com>
"""Data-formatting helpers for the MCP server."""

import logging

_logger = logging.getLogger(__name__)


def format_currency(value):
    """Formatea un valor como moneda"""
    if value is None:
        return "-"
    return f"{value:,.2f} €"


def format_percentage(value):
    """Formatea un valor como porcentaje"""
    if value is None:
        return "-"
    return f"{value:.2f}%"


def format_data_as_html_table(data, title, columns_config):
    """
    Formatea datos tabulares como HTML
    
    Args:
        data: Lista de diccionarios con los datos
        title: Título del reporte
        columns_config: Lista de tuplas (key, label, formatter, align)
            formatter puede ser: 'currency', 'percentage', 'number', 'text', None
            align puede ser: 'left', 'right', 'center'
    """
    from datetime import datetime
    
    generation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Generar filas
    rows_html = []
    for idx, row_data in enumerate(data, 1):
        cells = []
        for key, label, formatter, align in columns_config:
            value = row_data.get(key, '')
            
            # Aplicar formateo
            if formatter == 'currency':
                formatted_value = format_currency(value)
            elif formatter == 'percentage':
                formatted_value = format_percentage(value)
            elif formatter == 'number':
                formatted_value = f"{value:,.0f}" if value is not None else "-"
            else:
                formatted_value = str(value) if value is not None else "-"
            
            align_class = f"text-{align}" if align else ""
            cells.append(f'<td class="{align_class}">{formatted_value}</td>')
        
        row_html = f'                    <tr>\n                        <td class="text-center">{idx}</td>\n' + '\n                        '.join(cells) + '\n                    </tr>'
        rows_html.append(row_html)
    
    # Generar encabezados
    headers = ['<th>#</th>']
    for key, label, formatter, align in columns_config:
        align_class = f' class="text-{align}"' if align else ""
        headers.append(f'<th{align_class}>{label}</th>')
    
    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        .table-container {{
            padding: 30px;
            overflow-x: auto;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.95em;
        }}
        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.85em;
            letter-spacing: 0.5px;
        }}
        tbody tr {{
            border-bottom: 1px solid #e9ecef;
        }}
        tbody tr:hover {{
            background-color: #f8f9fa;
        }}
        tbody tr:nth-child(even) {{
            background-color: #fafbfc;
        }}
        td {{
            padding: 12px 15px;
        }}
        .text-right {{
            text-align: right;
        }}
        .text-center {{
            text-align: center;
        }}
        .footer {{
            padding: 20px 30px;
            background: #f8f9fa;
            text-align: center;
            color: #6c757d;
            font-size: 0.9em;
            border-top: 2px solid #e9ecef;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{title}</h1>
        </div>
        <div class="table-container">
            <table>
                <thead>
                    <tr>
{''.join(headers)}
                    </tr>
                </thead>
                <tbody>
{''.join(rows_html)}
                </tbody>
            </table>
        </div>
        <div class="footer">
            Generado el {generation_date} | Sistema MCP Odoo
        </div>
    </div>
</body>
</html>"""
    return html_content


def format_data_as_csv(data, columns_config):
    """
    Formatea datos tabulares como CSV
    
    Args:
        data: Lista de diccionarios con los datos
        columns_config: Lista de tuplas (key, label, formatter, align)
    """
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Encabezados
    headers = ['#']
    for key, label, formatter, align in columns_config:
        headers.append(label)
    writer.writerow(headers)
    
    # Datos
    for idx, row_data in enumerate(data, 1):
        row = [idx]
        for key, label, formatter, align in columns_config:
            value = row_data.get(key, '')
            
            # Aplicar formateo
            if formatter == 'currency':
                formatted_value = f"{value:,.2f}" if value is not None else "-"
            elif formatter == 'percentage':
                formatted_value = f"{value:.2f}%" if value is not None else "-"
            elif formatter == 'number':
                formatted_value = f"{value:,.0f}" if value is not None else "-"
            else:
                formatted_value = str(value) if value is not None else ""
            
            row.append(formatted_value)
        
        writer.writerow(row)
    
    return output.getvalue()


def format_data_as_xml(data, root_name, item_name, columns_config):
    """
    Formatea datos tabulares como XML
    
    Args:
        data: Lista de diccionarios con los datos
        root_name: Nombre del elemento raíz
        item_name: Nombre del elemento para cada fila
        columns_config: Lista de tuplas (key, label, formatter, align)
    """
    import xml.etree.ElementTree as ET
    from datetime import datetime
    
    # Crear elemento raíz
    root = ET.Element(root_name)
    root.set('generated', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    root.set('total_items', str(len(data)))
    
    # Añadir metadatos
    metadata = ET.SubElement(root, 'metadata')
    ET.SubElement(metadata, 'columns')
    for key, label, formatter, align in columns_config:
        col_elem = ET.SubElement(metadata.find('columns'), 'column')
        col_elem.set('key', key)
        col_elem.set('label', label)
        if formatter:
            col_elem.set('formatter', formatter)
        if align:
            col_elem.set('align', align)
    
    # Añadir datos
    items = ET.SubElement(root, 'items')
    for idx, row_data in enumerate(data, 1):
        item = ET.SubElement(items, item_name)
        item.set('index', str(idx))
        
        for key, label, formatter, align in columns_config:
            value = row_data.get(key, '')
            
            # Aplicar formateo
            if formatter == 'currency':
                formatted_value = f"{value:,.2f}" if value is not None else ""
            elif formatter == 'percentage':
                formatted_value = f"{value:.2f}" if value is not None else ""
            elif formatter == 'number':
                formatted_value = f"{value:,.0f}" if value is not None else ""
            else:
                formatted_value = str(value) if value is not None else ""
            
            # Usar nombre válido para XML (sin espacios ni caracteres especiales)
            xml_key = key.replace(' ', '_').replace('-', '_')
            field = ET.SubElement(item, xml_key)
            field.text = formatted_value
            field.set('label', label)
            if formatter:
                field.set('formatter', formatter)
            if align:
                field.set('align', align)
    
    # Convertir a string XML con indentación
    try:
        ET.indent(root, space="  ")
    except AttributeError:
        # ET.indent no disponible en Python < 3.9, usar alternativa
        pass
    
    xml_string = ET.tostring(root, encoding='unicode', xml_declaration=True)
    return xml_string


def format_data_as_excel(data, columns_config, title="Reporte",
                         include_row_index=True, caption=None, widths=None):
    """One-sheet wrapper around ``format_excel_workbook``."""
    return format_excel_workbook([{
        'data': data,
        'columns_config': columns_config,
        'title': title,
        'include_row_index': include_row_index,
        'caption': caption,
        'widths': widths,
    }])


def format_excel_workbook(sheets):
    """Styled .xlsx with one worksheet per spec (same look as a single sheet)."""
    try:
        from openpyxl import Workbook
        import io

        specs = [s for s in (sheets or []) if isinstance(s, dict)]
        if not specs:
            return None
        wb = Workbook()
        first = True
        for spec in specs:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            _fill_excel_sheet(ws, spec)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except ImportError:
        _logger.error("MCP: openpyxl no está instalado. Instala con: pip install openpyxl")
        raise Exception("openpyxl no está instalado. Instala con: pip install openpyxl")
    except Exception as e:
        _logger.error("MCP: Error generando Excel: %s", e)
        raise Exception("Error generando Excel: %s" % e)


def _fill_excel_sheet(ws, spec):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = spec.get('data') or []
    columns_config = spec.get('columns_config') or []
    title = spec.get('title') or 'Export'
    include_row_index = bool(spec.get('include_row_index'))
    caption = (spec.get('caption') or '').strip()
    widths = spec.get('widths')
    ws.title = str(title)[:31] or 'Export'

    header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    header_row = 2 if caption else 1
    data_col0 = 2 if include_row_index else 1
    n_data_cols = len(columns_config)
    n_cols = n_data_cols + (1 if include_row_index else 0)
    if caption and n_cols:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=max(1, n_cols),
        )
        cap = ws.cell(row=1, column=1, value=caption)
        cap.alignment = left_align
        cap.font = Font(bold=True, size=12)

    headers = []
    if include_row_index:
        headers.append('#')
    for key, label, formatter, align in columns_config:
        headers.append(label)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for row_offset, row_data in enumerate(data, 1):
        row_idx = header_row + row_offset
        if include_row_index:
            ws.cell(row=row_idx, column=1, value=row_offset).border = border
            ws.cell(row=row_idx, column=1).alignment = center_align

        for col_off, (key, label, formatter, align) in enumerate(columns_config):
            col_idx = data_col0 + col_off
            value = row_data.get(key, '') if isinstance(row_data, dict) else ''
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if formatter == 'currency':
                cell.value = float(value) if value is not None and value != '' else 0.0
                cell.number_format = '#,##0.00 €'
                cell.alignment = right_align
            elif formatter == 'percentage':
                if value is not None and value != '':
                    cell.value = float(value) / 100
                    cell.number_format = '0.00%'
                else:
                    cell.value = '-'
                cell.alignment = center_align
            elif formatter == 'number':
                cell.value = int(value) if value is not None and value != '' else 0
                cell.alignment = center_align
            else:
                if isinstance(value, bool):
                    cell.value = value
                    cell.alignment = center_align
                elif isinstance(value, int):
                    cell.value = value
                    cell.alignment = right_align
                elif isinstance(value, float):
                    if value != value or value in (float('inf'), float('-inf')):
                        cell.value = ''
                    else:
                        cell.value = value
                    cell.alignment = right_align
                else:
                    cell.value = str(value) if value is not None else ''
                    if align == 'center':
                        cell.alignment = center_align
                    elif align == 'right':
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align

    if include_row_index:
        ws.column_dimensions[get_column_letter(1)].width = 5
    for col_off, (key, label, formatter, align) in enumerate(columns_config):
        col_idx = data_col0 + col_off
        if widths and col_off < len(widths):
            width = widths[col_off]
        elif formatter == 'currency':
            width = 15
        elif formatter == 'percentage':
            width = 12
        elif formatter == 'number':
            width = 10
        else:
            width = max(len(str(label or '')), 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(float(width), 50)

    ws.freeze_panes = 'A%d' % (header_row + 1)



# Structural image-column hints (field shape, not a business glossary).
_PDF_IMG_KEY_HINTS = (
    'image', 'logo', 'avatar', 'photo',
    'map_osm', 'map_gm', 'mapthumb', 'staticmap',
)
_PDF_B64_IMG_PREFIXES = (
    ('iVBORw', 'png'),
    ('/9j/', 'jpeg'),
    ('R0lGOD', 'gif'),
    ('UklGR', 'webp'),
)
_PDF_IMG_SRC_RE = None
_PDF_WEB_IMAGE_RE = None
_PDF_DATA_URI_RE = None


def _pdf_regexes():
    import re
    global _PDF_IMG_SRC_RE, _PDF_WEB_IMAGE_RE, _PDF_DATA_URI_RE
    if _PDF_IMG_SRC_RE is None:
        _PDF_IMG_SRC_RE = re.compile(
            r'<img[^>]+src=["\']([^"\']+)["\']', re.I,
        )
        _PDF_WEB_IMAGE_RE = re.compile(
            r'/web/image/(?P<model>[\w.]+)/(?P<rid>\d+)/(?P<field>[\w.]+)',
        )
        _PDF_DATA_URI_RE = re.compile(
            r'^data:image/([^;]+);base64,(.+)$', re.I | re.DOTALL,
        )
    return _PDF_IMG_SRC_RE, _PDF_WEB_IMAGE_RE, _PDF_DATA_URI_RE


def _pdf_looks_image_key(key):
    key_l = str(key or '').lower()
    return any(hint in key_l for hint in _PDF_IMG_KEY_HINTS)


def _pdf_is_image_ref(value):
    """True when a cell is a photo reference, not printable text."""
    if value in (None, '', False, []):
        return False
    if isinstance(value, (bytes, bytearray)):
        return len(value) > 16
    text = str(value).strip()
    if text.startswith('data:image/') or '/web/image/' in text:
        return True
    for prefix, _mime in _PDF_B64_IMG_PREFIXES:
        if text.startswith(prefix) and len(text) > 64:
            return True
    return False


def _pdf_plain_cell(value, formatter):
    if value is None or value == '':
        return ''
    if formatter == 'currency':
        try:
            return f"{value:,.2f} €"
        except (TypeError, ValueError):
            return str(value)
    if formatter == 'percentage':
        try:
            return f"{value:.2f}%"
        except (TypeError, ValueError):
            return str(value)
    if formatter == 'number':
        try:
            return f"{value:,.0f}"
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def _pdf_decode_b64(raw):
    import base64
    text = raw.strip()
    if text.startswith(("b'", 'b"')) and text.endswith(("'", '"')):
        text = text[2:-1]
    try:
        return base64.b64decode(text, validate=False)
    except Exception:
        return None


def _pdf_image_from_env(path, env):
    """Bytes from /web/image/<model>/<id>/<field> using the user env."""
    if env is None or not path:
        return None
    _src, web_re, _data = _pdf_regexes()
    match = web_re.search(str(path))
    if not match:
        return None
    model = match.group('model')
    try:
        rec = env[model].browse(int(match.group('rid')))
        if not rec.exists():
            return None
        raw = rec[match.group('field')]
    except Exception:
        return None
    if not raw:
        return None
    if isinstance(raw, (bytes, bytearray)):
        data = bytes(raw)
        if data[:8] == b'\x89PNG\r\n\x1a\n' or data[:2] == b'\xff\xd8':
            return data
        decoded = _pdf_decode_b64(data.decode('ascii', errors='ignore'))
        return decoded or data
    if isinstance(raw, str):
        return _pdf_decode_b64(raw)
    return None


def _pdf_image_bytes(value, key=None, env=None):
    """Return image bytes for a cell, or None. Never emit a raw URL string."""
    if value in (None, '', False, []):
        return None
    src_re, web_re, data_re = _pdf_regexes()
    if isinstance(value, (bytes, bytearray)):
        raw = bytes(value)
        if len(raw) > 32 and (
            raw[:8] == b'\x89PNG\r\n\x1a\n' or raw[:2] == b'\xff\xd8'
        ):
            return raw
        decoded = _pdf_decode_b64(raw.decode('ascii', errors='ignore'))
        return decoded if decoded and len(decoded) > 32 else None
    text = str(value).strip()
    img_src = src_re.search(text)
    if img_src:
        text = img_src.group(1).strip()
    data_m = data_re.match(text)
    if data_m:
        return _pdf_decode_b64(data_m.group(2))
    for prefix, _mime in _PDF_B64_IMG_PREFIXES:
        if text.startswith(prefix) and len(text) > 64:
            return _pdf_decode_b64(text)
    if web_re.search(text) or text.startswith('/web/image/'):
        return _pdf_image_from_env(text, env)
    if _pdf_looks_image_key(key) and (
        text.startswith('http') or text.startswith('/')
    ):
        if web_re.search(text) or text.startswith('/web/image/'):
            return _pdf_image_from_env(text, env)
        return None
    return None


def _pdf_flowable_image(raw, max_side):
    from io import BytesIO
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image
    reader = ImageReader(BytesIO(raw))
    iw, ih = reader.getSize()
    if not iw or not ih:
        return None
    if iw >= ih:
        width = max_side
        height = max_side * float(ih) / float(iw)
    else:
        height = max_side
        width = max_side * float(iw) / float(ih)
    return Image(reader, width=width, height=height)


def _pdf_paragraph(text, style):
    from reportlab.platypus import Paragraph
    safe = (text or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    safe = safe.replace('\n', '<br/>')
    return Paragraph(safe or '&#160;', style)


_PDF_CHART_COLORS = (
    '#5b8def',
    '#26a69a',
    '#ef6c00',
    '#8d6e63',
)


def _pdf_chart_drawing(spec, width, height):
    """reportlab Drawing for a structural series. None if charts extras missing."""
    if not spec or not spec.get('categories') or not spec.get('series'):
        return None
    try:
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        from reportlab.graphics.charts.legends import Legend
        from reportlab.graphics.shapes import Drawing, String
        from reportlab.lib import colors
    except ImportError:
        return None
    cats = [str(cat)[:14] for cat in spec['categories']]
    series = spec['series']
    data = [tuple(item.get('values') or []) for item in series]
    if not data or not data[0]:
        return None
    kind = spec.get('kind') or ('line' if len(cats) >= 8 else 'bar')
    drawing = Drawing(width, height)
    left = 50
    bottom = 44
    right_pad = 14
    top_pad = 16
    legend_h = 14 if len(series) > 1 else 0
    chart_w = max(80, width - left - right_pad)
    chart_h = max(60, height - bottom - top_pad - legend_h)
    if kind == 'line':
        chart = HorizontalLineChart()
        chart.joinedLines = 1
    else:
        chart = VerticalBarChart()
        chart.barWidth = min(18, max(6, chart_w / max(len(cats) * len(series), 1) * 0.55))
        chart.groupSpacing = 6
        chart.barSpacing = 1
    chart.x = left
    chart.y = bottom
    chart.width = chart_w
    chart.height = chart_h
    chart.data = data
    chart.categoryAxis.categoryNames = cats
    chart.categoryAxis.labels.fontName = 'Helvetica'
    chart.categoryAxis.labels.fontSize = 6
    if len(cats) > 6:
        chart.categoryAxis.labels.angle = 40
        chart.categoryAxis.labels.boxAnchor = 'ne'
        chart.categoryAxis.labels.dy = -2
    else:
        chart.categoryAxis.labels.boxAnchor = 'n'
    chart.valueAxis.labels.fontName = 'Helvetica'
    chart.valueAxis.labels.fontSize = 7
    ymin = min(min(item['values']) for item in series)
    ymax = max(max(item['values']) for item in series)
    chart.valueAxis.valueMin = 0 if ymin >= 0 else ymin * 1.1
    chart.valueAxis.valueMax = ymax * 1.12 if ymax else 1
    for idx, hexcol in enumerate(_PDF_CHART_COLORS):
        color = colors.HexColor(hexcol)
        if kind == 'line':
            chart.lines[idx].strokeColor = color
            chart.lines[idx].strokeWidth = 1.6
        else:
            chart.bars[idx].fillColor = color
            chart.bars[idx].strokeColor = colors.white
            chart.bars[idx].strokeWidth = 0.3
    drawing.add(chart)
    if len(series) > 1:
        legend = Legend()
        legend.x = left
        legend.y = height - 10
        legend.fontName = 'Helvetica'
        legend.fontSize = 7
        legend.alignment = 'right'
        legend.colorNamePairs = [
            (
                colors.HexColor(_PDF_CHART_COLORS[idx % len(_PDF_CHART_COLORS)]),
                str(item.get('name') or '')[:24],
            )
            for idx, item in enumerate(series)
        ]
        drawing.add(legend)
    elif series:
        drawing.add(String(
            left, height - 10, str(series[0].get('name') or '')[:40],
            fontName='Helvetica', fontSize=7,
            fillColor=colors.HexColor('#5a5a5a'),
        ))
    return drawing


def format_data_as_pdf(data, columns_config, title="Reporte", landscape=True,
                       env=None, chart_series=None):
    """Tabular PDF with Chatboo table tokens (no row #, no MCP footer).

    Cells that are photos (data-URI, raw base64, /web/image via env) become
    thumbnails. Rows split between pages; a photo stays inside its row.
    A structural series (category + numbers) may prepend a reportlab chart.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape as a4_landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            KeepTogether, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
        import io

        buffer = io.BytesIO()
        pagesize = a4_landscape(A4) if landscape else A4
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1.4 * cm,
            bottomMargin=1 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ChatbooPdfTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=10,
            alignment=0,
        )
        cell_style = ParagraphStyle(
            'ChatbooPdfCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#282828'),
        )
        head_style = ParagraphStyle(
            'ChatbooPdfHead',
            parent=cell_style,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1a1a1a'),
        )

        story = []
        if title:
            story.append(_pdf_paragraph(str(title), title_style))
            story.append(Spacer(1, 0.2 * cm))
        if chart_series:
            try:
                usable_w = pagesize[0] - 2 * cm
                drawing = _pdf_chart_drawing(chart_series, usable_w, 7.2 * cm)
                if drawing is not None:
                    story.append(drawing)
                    story.append(Spacer(1, 0.35 * cm))
            except Exception:
                _logger.debug('format_data_as_pdf: chart drawing skipped', exc_info=True)

        header = []
        for key, label, formatter, align in columns_config:
            header.append(_pdf_paragraph(str(label or key), head_style))
        table_data = [header]
        thumb = 1.35 * cm
        image_rows = set()

        for row_idx, row_data in enumerate(data or []):
            row = []
            has_image = False
            for key, label, formatter, align in columns_config:
                value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                raw = _pdf_image_bytes(value, key=key, env=env)
                if raw:
                    flow = _pdf_flowable_image(raw, thumb)
                    if flow is not None:
                        row.append(KeepTogether([flow]))
                        has_image = True
                        continue
                if _pdf_is_image_ref(value):
                    row.append(_pdf_paragraph('', cell_style))
                    continue
                row.append(_pdf_paragraph(
                    _pdf_plain_cell(value, formatter), cell_style,
                ))
            table_data.append(row)
            if has_image:
                image_rows.add(row_idx + 1)

        row_heights = None
        if image_rows:
            row_heights = [None] * len(table_data)
            for idx in image_rows:
                row_heights[idx] = thumb + 0.25 * cm
        table = Table(
            table_data, repeatRows=1, splitByRow=1, rowHeights=row_heights,
        )
        table.hAlign = 'LEFT'
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eceff1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#282828')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#c8c8c8')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            (
                'ROWBACKGROUNDS', (0, 1), (-1, -1),
                [colors.white, colors.HexColor('#f8f9fa')],
            ),
        ]))
        for col_idx, (key, label, formatter, align) in enumerate(columns_config):
            if align == 'right':
                table.setStyle(TableStyle([
                    ('ALIGN', (col_idx, 1), (col_idx, -1), 'RIGHT'),
                ]))
            elif align == 'center':
                table.setStyle(TableStyle([
                    ('ALIGN', (col_idx, 1), (col_idx, -1), 'CENTER'),
                ]))
        story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    except ImportError:
        raise Exception("reportlab no está instalado. Instala con: pip install reportlab")
    except Exception as e:
        raise Exception(f"Error generando PDF: {str(e)}")

