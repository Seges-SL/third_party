# -*- coding: utf-8 -*-
# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025-2026 Patanegra Soft <https://patanegra.com>
"""PNS AI Chatboo - Async Request. PATANEGRA Soft (https://patanegra.com).

Part of Patanegra Soft Suite (`pns_suite`), distributed via Patanegra Soft Hub.
Asynchronous queue of chat turns for the Patanegra Application Agent Protocol
(PAAP). A turn runs in a thread with its own cursor (decoupled from the HTTP
connection), writing its progress and result to the database, saving the message
to the session and notifying over the bus. The response therefore survives an F5
or a tab close: the source of truth is the database, not the connection. The SSE
at /chatboo/stream only tails this job for the live preview.
Licensed under the Apache License 2.0 - see LICENSE.
"""

import base64
import json
import logging
import threading
import time
from datetime import timedelta

import odoo
from odoo import models, fields, api, _

from ..utils.turn_deliverable import done_meta_has_deliverable

_logger = logging.getLogger(__name__)

# Sentinela: "no tocar el estado del skill activo de la sesión" al finalizar
# (distinto de code=None, que significa soltarlo).
_SKILL_STATE_LEAVE = object()

# Intervalo de flush del progreso a BD (segundos). Granularidad del preview.
FLUSH_INTERVAL = 0.3
# Intervalo del LATIDO (heartbeat) independiente (segundos). Un hilo aparte
# refresca write_date a este ritmo mientras el job corre, PASE lo que pase con el
# motor: así un modelo lento que tarda minutos en emitir tokens no se confunde
# nunca con un worker muerto. Debe ser < HEARTBEAT_STALE_SECONDS con holgura.
HEARTBEAT_INTERVAL = 10.0
# Antigüedad para considerar un job "colgado" (min) → lo reclama el cron. Gracias
# al latido (write_date se refresca cada ~0.3s), un valor bajo es seguro: solo los
# jobs realmente muertos (sin latido reciente) superan el umbral.
STUCK_MINUTES = 3
# Umbral de latido para el reclaim inmediato al abrir Chatboo (segundos). Un job
# vivo late cada ~0.3s; si lleva más de esto sin latir, está muerto (reinicio).
HEARTBEAT_STALE_SECONDS = 60
# Retención de jobs terminados (horas) antes de purgarlos.
GC_HOURS = 72


def _env_manage():
    """Context manager de entornos thread-local.

    Odoo 14 (owl1) exige ``Environment.manage()`` para crear un Environment con
    cursor propio fuera del request. En Odoo 17+ (owl2) desapareció; devolvemos
    un nullcontext para que el mismo código sirva en ambos stacks.
    """
    manage = getattr(api.Environment, 'manage', None)
    if manage is not None:
        return manage()

    import contextlib
    return contextlib.nullcontext()


def _get_registry(dbname):
    """Devuelve el Registry de la BD, compatible con Odoo 14 y 17+/19.

    En Odoo 14 existe el helper ``odoo.registry(db)``. En Odoo 17+/19 se eliminó
    (``module 'odoo' has no attribute 'registry'``) y hay que usar la clase
    ``odoo.modules.registry.Registry(db)``.
    """
    reg = getattr(odoo, 'registry', None)
    if callable(reg):
        return reg(dbname)
    from odoo.modules.registry import Registry
    return Registry(dbname)


def _is_pg_serialization_failure(exc):
    """True for Postgres SQLSTATE 40001 (REPEATABLE READ concurrent update).

    Odoo 14/psycopg2 and Odoo 19/psycopg3 expose this under different class
    names; matching pgcode/sqlstate plus the well-known message covers both.
    """
    pgcode = getattr(exc, 'pgcode', None) or getattr(exc, 'sqlstate', None)
    if pgcode == '40001':
        return True
    if type(exc).__name__ == 'SerializationFailure':
        return True
    msg = str(exc or '').lower()
    return 'could not serialize access' in msg


def _job_already_finished(dbname, rid):
    """True si ``_finalize`` ya dejó el job en done/error (cursor fresco)."""
    try:
        with _get_registry(dbname).cursor() as cr:
            cr.execute(
                "SELECT state FROM chatboo_async_request WHERE id=%s",
                (rid,),
            )
            row = cr.fetchone()
            return bool(row and row[0] in ('done', 'error'))
    except Exception:
        return False


def _is_test_mode(env):
    """True si estamos en modo test, compatible con Odoo 14 y 17+/19.

    En Odoo 14 el registry expone ``in_test_mode()``. En Odoo 19 ese método ya no
    existe (``'Registry' object has no attribute 'in_test_mode'``), así que nos
    apoyamos en el flag ``testing`` del hilo y solo llamamos a ``in_test_mode`` si
    está disponible.
    """
    if getattr(threading.current_thread(), 'testing', False):
        return True
    fn = getattr(env.registry, 'in_test_mode', None)
    if callable(fn):
        try:
            return bool(fn())
        except Exception:
            return False
    return False


class ChatbooAsyncRequest(models.Model):
    """Turno de Chatboo ejecutado en segundo plano (desacoplado del request)."""
    _name = 'chatboo.async.request'
    _description = 'Chatboo Async Request'
    _order = 'create_date desc'
    _rec_name = 'id'

    session_id = fields.Many2one(
        'chatboo.session', string='Session', required=True,
        ondelete='cascade', index=True,
    )
    user_id = fields.Many2one(
        'res.users', string='User', required=True, index=True,
        default=lambda self: self.env.user,
    )
    prompt = fields.Text(string='Prompt', required=True)
    history = fields.Text(string='History (JSON)')
    agent_code = fields.Char(string='Agent code')
    provider_id = fields.Integer(string='Provider id')
    # Descriptor ligero capturado en el cliente (modelo, id, vista) en JSON.
    screen_context = fields.Text(string='Screen context (JSON)')
    # Imágenes del turno (data URLs, JSON list). El base64 viaja al proveedor de
    # visión SOLO en este turno y se purga con el job (GC 72h); NUNCA entra en el
    # historial que va al LLM (doctrina anti-overflow). Para MOSTRAR, en cambio,
    # sí se persisten una vez como ir.attachment ligado a la sesión y el mensaje
    # de usuario guarda solo su URL (ver _persist_turn_images).
    images = fields.Text(string='Images (JSON)')
    # Nombres originales de las imágenes del turno, alineados por índice con
    # `images` (JSON list). null para las pegadas del portapapeles (sin nombre)
    # y el nombre real para las adjuntadas como fichero. Se usa para nombrar el
    # ir.attachment y mostrar el nombre bajo la miniatura (ver _persist_turn_images).
    image_names = fields.Text(string='Image names (JSON)')
    # Ficheros de texto/datos adjuntados con el clip (JSON list de
    # {name, mimetype, data(dataURL base64)}). Su TEXTO extraído se inyecta en el
    # mensaje de ESTE turno hacia el LLM (transitorio, no entra en el historial);
    # el fichero se persiste como ir.attachment y el mensaje guarda un chip
    # descargable (ver _persist_turn_files / _extract_file_text).
    files = fields.Text(string='Files (JSON)')

    state = fields.Selection(
        [
            ('pending', 'Pending'),
            ('running', 'Running'),
            ('done', 'Done'),
            ('error', 'Error'),
        ],
        string='State', default='pending', required=True, index=True,
    )
    # Contenido visible acumulado (tokens/replace) para el preview y el guardado.
    partial = fields.Text(string='Partial content')
    # Eventos estructurados no-token (status, verification, ...) que el tail
    # reemite tal cual por SSE. JSON list.
    struct_events = fields.Text(string='Structured events (JSON)')
    # Metadatos del evento 'done' del motor (usage, context_limit, speed, ...).
    done_meta = fields.Text(string='Done meta (JSON)')
    response = fields.Text(string='Final response')
    error = fields.Text(string='Error')
    # True cuando un cliente ya recibió el resultado en vivo (SSE) o por poll.
    seen = fields.Boolean(string='Seen by client', default=False, index=True)
    # El cliente pidió cancelar: el worker lo comprueba entre eventos y cierra.
    cancel_requested = fields.Boolean(
        string='Cancel requested', default=False, index=True,
    )

    started_at = fields.Datetime(string='Started at')
    finished_at = fields.Datetime(string='Finished at')

    # ──────────────────────────── API pública ────────────────────────────

    @api.model
    def enqueue(self, session_id, message, history=None, agent_code=None,
                provider_id=None, screen_context=None, images=None, files=None,
                image_names=None):
        """Crea el job en estado pending. No lanza el hilo (ver ``spawn``)."""
        sc_json = False
        if screen_context:
            if isinstance(screen_context, str):
                sc_json = screen_context
            else:
                sc_json = json.dumps(screen_context, ensure_ascii=False)
        return self.create({
            'session_id': int(session_id),
            'user_id': self.env.uid,
            'prompt': message or '',
            'history': json.dumps(history or [], ensure_ascii=False),
            'agent_code': agent_code or False,
            'provider_id': int(provider_id) if provider_id else False,
            'screen_context': sc_json,
            'images': json.dumps(images, ensure_ascii=False) if images else False,
            'image_names': json.dumps(image_names, ensure_ascii=False) if image_names else False,
            'files': json.dumps(files, ensure_ascii=False) if files else False,
            'state': 'pending',
        })

    def spawn(self):
        """Lanza el procesamiento en un hilo con cursor propio.

        En modo test se ejecuta en línea (los hilos con cursor propio y el
        TestCursor no se llevan bien).
        """
        self.ensure_one()
        rid = self.id
        dbname = self.env.cr.dbname
        uid = self.env.uid
        context = dict(self.env.context)

        if _is_test_mode(self.env):
            self._process()
            return

        def _runner():
            # Patrón estándar de hilo de fondo Odoo: fijar dbname en el hilo (si
            # no, sale "desnudo" -> el log muestra '?' y algunas rutas del ORM/log
            # que lo esperan pueden fallar).
            try:
                threading.current_thread().dbname = dbname
            except Exception:
                pass
            _logger.info("Chatboo async worker thread started (request %s, db %s)", rid, dbname)
            try:
                with _env_manage():
                    db_registry = _get_registry(dbname)
                    with db_registry.cursor() as cr:
                        env = api.Environment(cr, uid, context)
                        env['chatboo.async.request'].browse(rid)._process()
            except Exception as exc:
                # Commit del cursor del worker al salir del ``with``: un
                # SerializationFailure (PG 40001) aquí NO es un fallo del
                # motor — _finalize ya persistió en un cursor fresco. Pintar
                # "MCP engine error" pisaría un turno bueno (tabla/texto ya
                # emitidos por SSE).
                if _is_pg_serialization_failure(exc):
                    _logger.warning(
                        "Chatboo async worker cursor raced after finalize "
                        "(request %s): %s",
                        rid, exc,
                    )
                    if _job_already_finished(dbname, rid):
                        return
                _logger.exception("Chatboo async worker crashed (request %s)", rid)
                self._record_crash(dbname, uid, context, rid, exc)

        thread = threading.Thread(
            target=_runner, name='chatboo-async-%s' % rid, daemon=True,
        )
        thread.start()

    @api.model
    def _record_crash(self, dbname, uid, context, rid, exc):
        """Marca un job como error tras un crash de bootstrap del hilo.

        Intenta primero la vía ORM (guarda en sesión + avisa por bus); si el
        propio Environment es lo que falla, cae a SQL crudo para al menos dejar el
        estado en 'error' con el mensaje, que el tail SSE ya reemite al cliente.
        """
        if _job_already_finished(dbname, rid):
            _logger.warning(
                "Chatboo async: skip crash mark, job %s already finished", rid,
            )
            return
        msg = '%s: %s' % (exc.__class__.__name__, exc) or 'worker crashed'
        try:
            # _env_manage() es imprescindible en Odoo 14 para crear un Environment
            # fuera del request (si no: 'tuple' object has no attribute 'cache').
            with _env_manage():
                with _get_registry(dbname).cursor() as cr:
                    env = api.Environment(cr, uid, context)
                    env['chatboo.async.request'].browse(rid)._finalize('', [], {}, error=msg)
                    return
        except Exception:
            _logger.exception("Chatboo async: _finalize on crash failed (request %s)", rid)
        try:
            with _get_registry(dbname).cursor() as cr:
                cr.execute(
                    "UPDATE chatboo_async_request "
                    "SET state='error', error=%s, response='', "
                    "finished_at=(now() at time zone 'UTC') WHERE id=%s",
                    (msg, rid),
                )
        except Exception:
            _logger.exception("Chatboo async: raw crash mark failed (request %s)", rid)

    # ──────────────────────────── Ejecución ────────────────────────────

    def _process(self):
        """Ejecuta el motor, va volcando progreso a BD, guarda en sesión y avisa."""
        self.ensure_one()
        _logger.info(
            "Chatboo async job %s: _process start (agent=%s, provider=%s)",
            self.id, self.agent_code or '-', self.provider_id or '-',
        )
        self.write({'state': 'running', 'started_at': fields.Datetime.now()})
        self.env.cr.commit()
        self._notify('thinking')

        acc = ''
        struct = []
        done_meta = {}
        last_flush = 0.0
        # Ancla de reutilización de consulta (Nivel 1): recogemos el código de la
        # última consulta exitosa que emita el motor (evento 'query_code') para
        # guardarlo en la sesión, y leemos el del turno anterior para que el
        # motor lo reinyecte como pista (evita que reescriba la query en cada
        # reformateo → fuente principal de variabilidad).
        last_query_code = None
        last_query_data = None
        try:
            prior_query_code = self.session_id.last_query_code or None
        except Exception:
            prior_query_code = None
        # Nivel 2: dataset del turno anterior SOLO si sigue fresco (TTL). Se
        # reinyecta como previous_result para reformatear sin re-consultar.
        try:
            prior_query_data = self.session_id.get_fresh_query_data()
        except Exception:
            prior_query_data = None
        # Skill pegajoso: código + params del skill activo de la sesión. El motor
        # los usa para juzgar (vía IA) si este turno refina el skill activo.
        try:
            prior_active_skill_code = self.session_id.active_skill_code or None
            prior_active_skill_params = self.session_id.active_skill_params or None
        except Exception:
            prior_active_skill_code = None
            prior_active_skill_params = None
        # Layout showmode: NL phrasing. Slash /show-* is handled later (writes session).
        try:
            _raw_prompt = (self.prompt or '').strip().lower()
            if not (
                _raw_prompt.startswith('/show-table')
                or _raw_prompt.startswith('/show-chart')
            ):
                from odoo.addons.pns_ai_mcp.utils.presentation_mode import (
                    apply_sticky_show_mode,
                )
                apply_sticky_show_mode(
                    self.session_id,
                    self.prompt or '',
                    user_lang=self.env.user.lang,
                )
        except Exception:
            _logger.debug(
                'Chatboo async job %s: showmode sticky skipped',
                self.id, exc_info=True,
            )
        # Estado del skill activo emitido por el motor este turno (sentinela =
        # no tocado → no reescribir la sesión).
        active_skill_touched = False
        active_skill_evt = None

        try:
            history = json.loads(self.history) if self.history else []
        except Exception:
            history = []

        # Imágenes del turno (data URLs). Solo para el turno actual: van al
        # proveedor de visión como content-parts; no entran en el historial.
        try:
            images = json.loads(self.images) if self.images else []
        except Exception:
            images = []

        # Ficheros de texto/datos adjuntos con el clip: extraemos su texto y lo
        # inyectamos en el mensaje de ESTE turno (no en self.prompt ni en el
        # historial). Así el LLM ve el contenido del fichero solo cuando se
        # adjunta, igual que las imágenes (transitorio, anti-overflow).
        message_for_engine = self._augment_prompt_with_files(self.prompt or '')

        # Axis slashes: one-shot painter/footmode; showmode writes the session.
        formatting_override = None
        turn_painter = None
        turn_footmode = None
        try:
            from odoo.addons.pns_ai_mcp.utils.formatting_mode_policy import (
                AXIS_FOOTMODE,
                AXIS_PAINTER,
                apply_turn_axis,
                resolve_remote_formatting_override,
            )
            axis, value, stripped = apply_turn_axis(
                self.session_id,
                message_for_engine,
            )
            if axis is not None:
                message_for_engine = stripped
                if not (message_for_engine or '').strip():
                    message_for_engine = '/' + value
                if axis == AXIS_PAINTER:
                    turn_painter = value
                elif axis == AXIS_FOOTMODE:
                    turn_footmode = value
            formatting_override = resolve_remote_formatting_override(
                turn_mode=turn_painter,
            )
        except Exception:
            formatting_override = None
            _logger.debug(
                'Chatboo async job %s: presentation axis turn parse skipped',
                self.id, exc_info=True,
            )
        # Recuerdo acotado de imágenes previas (1c): reenviamos, con tope y
        # reescaladas, las últimas imágenes ya vistas en la conversación para
        # que el modelo pueda referirse a ellas (los LLM son sin estado: la
        # imagen se "recuerda" solo si se REENVÍA). No entran en el historial
        # persistido; se recomponen cada turno desde los ir.attachment.
        #
        # PERO: si ESTE turno YA trae imágenes, NO recordamos las anteriores.
        # Cuando el usuario adjunta imágenes y pregunta "¿qué hay en estas?", se
        # refiere a las del prompt ACTUAL; reenviar imágenes viejas hacía que el
        # modelo mezclara/mencionara una imagen de hace varios turnos. El recuerdo
        # queda para los turnos de SOLO texto que aludan a algo ya enviado.
        recall_images = [] if images else self._recall_prior_images()

        # LATIDO independiente: un hilo aparte refresca write_date cada
        # HEARTBEAT_INTERVAL mientras dura el turno, PASE lo que pase con el motor.
        # Así un modelo lento (que puede tardar minutos en emitir el primer token)
        # no se confunde nunca con un worker muerto y el reclaim no lo mata.
        beat = self._start_heartbeat()

        # Diagnóstico: contamos qué emite el motor. Un turno que acaba en 0s y en
        # blanco casi siempre significa que el motor no llegó a emitir tokens
        # (proveedor no resuelto en el hilo, endpoint vacío, etc.). Este desglose
        # en el log es lo que permite distinguir "sin contenido" de "crash".
        ev_counts = {}
        screen_context_block = ''
        try:
            raw_sc = json.loads(self.screen_context) if self.screen_context else None
            if raw_sc:
                from odoo.addons.pns_ai_chatboo.utils.screen_context import enrich_screen_context
                enriched = enrich_screen_context(self.env, raw_sc)
                screen_context_block = enriched.get('block') or ''
                if screen_context_block:
                    _logger.info(
                        'Chatboo async job %s: screen_context enriched (%d chars)',
                        self.id, len(screen_context_block),
                    )
                if self.session_id:
                    try:
                        self.session_id.sudo().write({
                            'last_screen_context': json.dumps(
                                raw_sc, ensure_ascii=False,
                            ),
                        })
                    except Exception:
                        _logger.debug(
                            'Chatboo async job %s: could not store last screen',
                            self.id, exc_info=True,
                        )
        except Exception:
            _logger.warning(
                'Chatboo async job %s: screen_context enrich failed',
                self.id, exc_info=True,
            )

        try:
            from odoo.addons.pns_ai_mcp.utils.agent_engine import AgentEngine
            engine_env = self.env(context=dict(
                self.env.context or {},
                chatboo_session_id=self.session_id.id,
                user_message=self.prompt or '',
                llm_remote_formatting_override=formatting_override,
                turn_painter=turn_painter,
                turn_footmode=turn_footmode,
            ))
            engine = AgentEngine(engine_env)
            for evt in engine.run_stream(
                message_for_engine, history, self.session_id.id,
                agent_code=self.agent_code or None,
                consumer_key='chatboo',
                provider_id=self.provider_id or None,
                prior_query_code=prior_query_code,
                prior_query_data=prior_query_data,
                screen_context_block=screen_context_block or None,
                prior_active_skill_code=prior_active_skill_code,
                prior_active_skill_params=prior_active_skill_params,
                images=images or None,
                recall_images=recall_images or None,
            ):
                if self._is_cancel_requested():
                    _logger.info("Chatboo async job %s: cancel requested, stopping", self.id)
                    break
                ev = (evt or {}).get('event', 'token')
                ev_counts[ev] = ev_counts.get(ev, 0) + 1
                if ev == 'active_skill':
                    # Meta-evento interno: estado del skill pegajoso (code+params,
                    # o code=None para soltarlo). Se persiste al finalizar.
                    active_skill_touched = True
                    active_skill_evt = evt
                    continue
                if ev == 'query_code':
                    # Meta-evento interno: no es contenido ni un evento de tail;
                    # solo memorizamos el último código exitoso del turno.
                    if evt.get('code'):
                        last_query_code = evt['code']
                    continue
                if ev == 'query_data':
                    # Meta-evento interno (Nivel 2): filas del último dataset
                    # exitoso, para cachearlas y reutilizarlas el próximo turno.
                    if evt.get('data') is not None:
                        last_query_data = evt['data']
                    continue
                if ev == 'token' and evt.get('content'):
                    acc += evt['content']
                elif ev == 'replace':
                    # 'replace' fija el contenido acumulado. Un replace VACÍO es
                    # una señal de reset: el motor lo emite al hacer failover tras
                    # haber emitido parcialmente, para que no se DUPLIQUE la
                    # respuesta al regenerarla el siguiente proveedor.
                    acc = evt.get('content') or ''
                elif ev == 'done':
                    done_meta = evt
                elif ev == 'error':
                    if evt.get('content'):
                        acc += evt['content']
                    struct.append(evt)
                else:
                    # status, verification, ... → reemitir tal cual en el tail
                    struct.append(evt)

                now = time.time()
                if now - last_flush >= FLUSH_INTERVAL:
                    self._flush_progress(acc, struct)
                    last_flush = now
                    if self._is_cancel_requested():
                        _logger.info(
                            "Chatboo async job %s: cancel requested after flush",
                            self.id,
                        )
                        break

            cancelled = self._is_cancel_requested()
            if cancelled:
                note = '⏹ ' + _('Generation cancelled.')
                if (acc or '').strip():
                    acc = (acc.rstrip() + '\n\n' + note)
                else:
                    acc = note
                _logger.info("Chatboo async job %s: cancelled by user", self.id)

            _logger.info(
                "Chatboo async job %s: engine done. events=%s acc_len=%s agent=%s provider_id=%s",
                self.id, ev_counts, len(acc), self.agent_code or '(auto)', self.provider_id or '(auto)',
            )
            # Parar y ESPERAR el latido ANTES de persistir: si sigue vivo, su
            # update concurrente a la fila colisiona con _finalize.
            self._stop_heartbeat(beat)
            beat = None
            self._finalize(
                acc, struct, done_meta, error=None,
                query_code=last_query_code, query_data=last_query_data,
                active_skill=(active_skill_evt if active_skill_touched
                              else _SKILL_STATE_LEAVE),
            )
        except Exception as exc:  # noqa: BLE001 — el motor puede fallar de mil formas
            _logger.exception(
                "Chatboo async job %s failed (events so far=%s, acc_len=%s)",
                self.id, ev_counts, len(acc),
            )
            self._stop_heartbeat(beat)
            beat = None
            self._finalize(acc, struct, done_meta, error=str(exc) or exc.__class__.__name__)
        finally:
            # Parar el latido pase lo que pase (éxito o error).
            self._stop_heartbeat(beat)
            # Este cursor lleva un snapshot REPEATABLE READ de TODO el turno.
            # Lo durable (sesión, job, logs, recetas, usage) ya se commitó en
            # cursores propios. Un commit aquí choca con el latido/flush
            # (PG 40001) y spawn() pintaba "MCP engine error" sobre un turno
            # bueno. En tests el TestCursor ES la transacción del caso: no
            # rollback.
            if not _is_test_mode(self.env):
                try:
                    self.env.cr.rollback()
                except Exception:
                    pass

    def _start_heartbeat(self):
        """Lanza el hilo de latido y devuelve ``(Event, Thread)`` para pararlo.

        Refresca ``write_date`` cada ``HEARTBEAT_INTERVAL`` en un cursor propio,
        sin tocar el contenido. En modo test no se lanza (el TestCursor y los
        hilos con cursor propio no se llevan bien): devuelve ``None``.
        """
        if _is_test_mode(self.env):
            return None
        stop_beat = threading.Event()
        rid = self.id
        dbname = self.env.cr.dbname

        def _beat():
            while not stop_beat.wait(HEARTBEAT_INTERVAL):
                try:
                    with _get_registry(dbname).cursor() as cr:
                        cr.execute(
                            "UPDATE chatboo_async_request "
                            "SET write_date=(now() at time zone 'UTC') WHERE id=%s",
                            (rid,),
                        )
                except Exception:
                    _logger.debug("Chatboo async heartbeat failed (request %s)", rid, exc_info=True)

        thread = threading.Thread(
            target=_beat, name='chatboo-beat-%s' % rid, daemon=True,
        )
        thread.start()
        return (stop_beat, thread)

    @staticmethod
    def _stop_heartbeat(beat):
        """Detiene el latido y ESPERA a que termine, para que no escriba la fila
        mientras ``_finalize`` la persiste (evita colisión de update concurrente)."""
        if not beat:
            return
        stop_beat, thread = beat
        try:
            stop_beat.set()
            thread.join(timeout=HEARTBEAT_INTERVAL + 2)
        except Exception:
            pass

    def _flush_progress(self, acc, struct):
        """Vuelca progreso parcial a BD en un cursor propio (visible para el tail).

        Refresca también ``write_date`` (latido): un job que sigue vivo se
        distingue así de uno muerto por un reinicio, que deja de latir.
        """
        try:
            with _get_registry(self.env.cr.dbname).cursor() as cr:
                cr.execute(
                    "UPDATE chatboo_async_request "
                    "SET partial=%s, struct_events=%s, write_date=(now() at time zone 'UTC') "
                    "WHERE id=%s",
                    (acc or '', json.dumps(struct, ensure_ascii=False), self.id),
                )
        except Exception:
            _logger.debug("Chatboo async flush failed (request %s)", self.id, exc_info=True)

    def _finalize(self, acc, struct, done_meta, error, query_code=None, query_data=None, active_skill=_SKILL_STATE_LEAVE):
        """Escribe estado final, persiste el turno en la sesión y avisa por bus."""
        # Red de seguridad anti-blanco: si el motor terminó "bien" pero sin
        # contenido, es un fallo real (proveedor no resuelto en el hilo, stream
        # vacío, etc.). Nunca dejamos un turno en blanco silencioso: lo marcamos
        # como error con un mensaje claro para que el usuario lo vea y se registre.
        # Excepción: exportación nominada (PDF/Excel/…) con tabla oculta a
        # propósito — el globo va vacío y la tarjeta/clip_data son el resultado
        # (C8N1: pintar "MCP engine error" encima de un PDF bueno).
        if not error and not (acc or '').strip():
            if done_meta_has_deliverable(done_meta):
                _logger.info(
                    "Chatboo async job %s finished with empty bubble and a "
                    "session file/clip_data; treating as success.",
                    self.id,
                )
            else:
                _logger.warning(
                    "Chatboo async job %s finished with EMPTY content and no error "
                    "(done_meta=%s). Marking as error to avoid a blank turn.",
                    self.id, done_meta,
                )
                error = _('The model returned no content. Check the AI provider '
                          '(endpoint reachable, model loaded) and the server log.')
        vals = {
            'partial': acc or '',
            'struct_events': json.dumps(struct, ensure_ascii=False),
            'done_meta': json.dumps(done_meta or {}, ensure_ascii=False),
            'response': acc or '',
            'finished_at': fields.Datetime.now(),
        }
        if error:
            vals['state'] = 'error'
            vals['error'] = error
        else:
            vals['state'] = 'done'
            vals['error'] = False

        # Persistencia en una transacción FRESCA (cursor propio), NO en el cursor
        # principal del worker: ese cursor arrastra un snapshot viejo y las
        # escrituras concurrentes a esta misma fila (flush cada 0.3s, latido)
        # hacen que su commit final falle con "could not serialize access due to
        # concurrent update" (REPEATABLE READ), abortando la transacción y dejando
        # el turno en rojo pese a haberse generado. Un cursor nuevo ve los últimos
        # commits y escribe sin colisión.
        dbname = self.env.cr.dbname
        uid = self.env.uid
        context = dict(self.env.context)
        rid = self.id
        try:
            with _env_manage():
                with _get_registry(dbname).cursor() as cr:
                    env = api.Environment(cr, uid, context)
                    job = env['chatboo.async.request'].browse(rid)
                    job.write(vals)
                    try:
                        _chips = job._save_to_session(acc, done_meta, error, query_code=query_code, query_data=query_data, active_skill=active_skill)
                        # Propagar los chips persistidos del turno de usuario al
                        # done_meta (ya escrito en vals): el controlador vuelca
                        # done_meta en el evento 'done' y el cliente parchea la
                        # burbuja en vivo (base64 → /web/image, ficheros clicables).
                        if _chips and (
                            _chips.get('user_images')
                            or _chips.get('user_files')
                            or _chips.get('assistant_files')
                            or _chips.get('assistant_content') is not None
                        ):
                            _merged = dict(done_meta or {})
                            _merged.update(_chips)
                            _write = {'done_meta': json.dumps(_merged, ensure_ascii=False)}
                            if _chips.get('assistant_content') is not None:
                                _write['response'] = _chips['assistant_content']
                                _write['partial'] = _chips['assistant_content']
                            job.write(_write)
                    except Exception:
                        _logger.exception(
                            "Chatboo async: failed to persist turn to session (job %s)", rid)
                    cr.commit()
                    job._notify('error' if error else 'async_done')
            return
        except Exception:
            _logger.exception("Chatboo async: _finalize fresh-cursor failed (job %s)", rid)

        # Último recurso: dejar el estado final por SQL crudo para no colgar el job
        # (el tail SSE reemite el estado/error al cliente aunque no haya sesión).
        try:
            with _get_registry(dbname).cursor() as cr:
                cr.execute(
                    "UPDATE chatboo_async_request "
                    "SET state=%s, error=%s, response=%s, partial=%s, "
                    "finished_at=(now() at time zone 'UTC') WHERE id=%s",
                    (vals['state'], error or None, acc or '', acc or '', rid),
                )
        except Exception:
            _logger.exception("Chatboo async: _finalize raw fallback failed (job %s)", rid)

    # ── Persistencia de imágenes del turno (para MOSTRAR, no para el LLM) ──────
    _IMG_EXT_BY_MIME = {
        'image/png': '.png', 'image/jpeg': '.jpg', 'image/jpg': '.jpg',
        'image/gif': '.gif', 'image/webp': '.webp', 'image/bmp': '.bmp',
        'image/svg+xml': '.svg', 'image/x-icon': '.ico',
        'image/vnd.microsoft.icon': '.ico',
    }

    @staticmethod
    def _split_data_url(data_url):
        """Parte un data URL 'data:<mime>;base64,<b64>' en (mimetype, b64).

        Si llega solo el base64 (sin prefijo), asume PNG.
        """
        s = data_url or ''
        if s.startswith('data:'):
            try:
                head, b64 = s.split(',', 1)
            except ValueError:
                return 'image/png', ''
            mimetype = head[5:].split(';', 1)[0] or 'image/png'
            return mimetype, b64
        return 'image/png', s

    def _persist_turn_images(self):
        """Persiste las imágenes del turno como ir.attachment y devuelve chips.

        Doctrina revisada: el base64 NUNCA entra en el blob JSON de la sesión ni
        en el historial que viaja al LLM. Pero SÍ persistimos cada imagen una vez
        como adjunto ligado a la sesión, y el mensaje de usuario guarda un chip
        ``{url, name}`` (/web/image con access_token, robusto en O14 y O19). Así el
        histórico conserva las capturas (multi-dispositivo), se abren en pestaña y
        conservan su NOMBRE cuando se adjuntaron como fichero (las pegadas del
        portapapeles no tienen nombre → name=None). Al borrar la sesión, su unlink
        arrastra los ir.attachment ligados (res_model/res_id), así que no se acumulan.
        """
        try:
            imgs = json.loads(self.images) if self.images else []
        except Exception:
            imgs = []
        try:
            names = json.loads(self.image_names) if self.image_names else []
        except Exception:
            names = []
        out = []
        if not imgs:
            return out
        Attachment = self.env['ir.attachment']
        for idx, data_url in enumerate(imgs):
            if not data_url:
                continue
            mimetype, b64 = self._split_data_url(data_url)
            if not b64:
                continue
            ext = self._IMG_EXT_BY_MIME.get((mimetype or '').lower(), '.png')
            orig = names[idx] if idx < len(names) else None
            orig = (orig or '').strip() or None
            att_name = orig or ('chatboo-%s-%s%s' % (self.id, idx + 1, ext))
            try:
                att = Attachment.create({
                    'name': att_name,
                    'datas': b64,
                    'mimetype': mimetype or 'image/png',
                    'res_model': 'chatboo.session',
                    'res_id': self.session_id.id,
                })
                token = att.generate_access_token()[0]
                out.append({
                    'url': '/web/image/%s?access_token=%s' % (att.id, token),
                    'name': orig,  # None para pegadas del portapapeles
                })
            except Exception:
                _logger.warning(
                    'Chatboo async job %s: no se pudo persistir la imagen %s del turno',
                    self.id, idx, exc_info=True,
                )
        return out

    # ── Recuerdo acotado de imágenes previas (1c) ─────────────────────────────
    # Cuántas imágenes previas de la sesión se reenvían como contexto y a qué
    # lado máximo se reescalan (para acotar tokens, como hacen ChatGPT/Claude).
    _RECALL_MAX_IMAGES = 3
    _RECALL_MAX_SIDE = 768
    # Formatos RÁSTER que los modelos de visión saben decodificar. Se EXCLUYE
    # image/svg+xml (y otros vectoriales/iconos): un SVG en el recall hacía que
    # TODOS los proveedores rechazaran el turno ("cannot identify image file" /
    # "does not contain a valid JPG, PNG, WebP, or ICO image") y disparaba el
    # failover en cascada hasta agotarlos. Además, la sesión acumula adjuntos que
    # no son fotos del usuario (iconos de skills como pin.svg, thumbnails de
    # registros…); filtrar por estos MIME evita reenviar basura vectorial.
    _RECALL_VISION_MIMETYPES = (
        'image/jpeg', 'image/jpg', 'image/png',
        'image/gif', 'image/webp', 'image/bmp',
    )

    def _recall_prior_images(self):
        """Data URLs (reescaladas) de las últimas imágenes ya vistas en la sesión.

        Reenvía, ACOTADO, las imágenes de turnos anteriores para que el modelo
        pueda referirse a ellas (los LLM son sin estado: solo "recuerdan" lo que
        se reenvía). Las del turno ACTUAL aún no están persistidas aquí, así que
        no se duplican. Devuelve orden cronológico.
        """
        if not self.session_id:
            return []
        try:
            # Pedimos más de la cuenta para poder DEDUPLICAR por checksum: la misma
            # foto adjuntada en varios turnos crea varios ir.attachment idénticos;
            # reenviarla repetida gasta contexto y confunde al modelo.
            atts = self.env['ir.attachment'].search([
                ('res_model', '=', 'chatboo.session'),
                ('res_id', '=', self.session_id.id),
                ('mimetype', 'in', list(self._RECALL_VISION_MIMETYPES)),
            ], order='id desc', limit=self._RECALL_MAX_IMAGES * 3)
        except Exception:
            return []
        urls = []
        seen = set()
        for att in atts:
            try:
                key = att.checksum or att.id
                if key in seen:
                    continue
                seen.add(key)
                raw = base64.b64decode(att.datas or b'')
                if not raw:
                    continue
                data_url = self._downscale_image_to_data_url(raw, att.mimetype)
                if data_url:
                    urls.append(data_url)
                if len(urls) >= self._RECALL_MAX_IMAGES:
                    break
            except Exception:
                continue
        urls.reverse()  # de más antigua a más reciente
        return urls

    @classmethod
    def _downscale_image_to_data_url(cls, raw, mimetype):
        """Reescala una imagen al lado máx. y la devuelve como data URL base64.

        Usa Pillow (dependencia de Odoo). Si falla (formato raro, sin Pillow),
        devuelve la imagen original en base64: el tope de NÚMERO ya acota el coste.
        """
        try:
            import io
            from PIL import Image
            im = Image.open(io.BytesIO(raw))
            im.load()
            src_fmt = (im.format or 'PNG').upper()
            w, h = im.size
            scale = cls._RECALL_MAX_SIDE / float(max(w, h) or 1)
            if scale < 1.0:
                im = im.resize((max(1, int(w * scale)), max(1, int(h * scale))))
            save_fmt = 'JPEG' if src_fmt in ('JPEG', 'JPG') else 'PNG'
            if save_fmt == 'JPEG' and im.mode not in ('RGB', 'L'):
                im = im.convert('RGB')
            out = io.BytesIO()
            im.save(out, format=save_fmt)
            b64 = base64.b64encode(out.getvalue()).decode('ascii')
            mt = 'image/jpeg' if save_fmt == 'JPEG' else 'image/png'
            return 'data:%s;base64,%s' % (mt, b64)
        except Exception:
            # Fallback SIN Pillow (o formato que no abre): reenviamos el original
            # SOLO si es un ráster que la visión acepta. NUNCA emitir SVG u otros
            # vectoriales/desconocidos: los proveedores los rechazan y tumban el
            # turno con failover en cascada.
            if (mimetype or '').lower() not in cls._RECALL_VISION_MIMETYPES:
                return ''
            try:
                b64 = base64.b64encode(raw).decode('ascii')
                return 'data:%s;base64,%s' % (mimetype, b64)
            except Exception:
                return ''

    # ── Adjuntos de texto/datos del clip (Fase 1) ─────────────────────────────
    # Política explícita: NUNCA truncar el contenido extraído de un adjunto.
    # El modelo debe ver el fichero completo (todas las pestañas Excel incluidas).

    @staticmethod
    def _decode_text_bytes(raw):
        """Decodifica bytes a texto probando utf-8/BOM y cayendo a latin-1/replace."""
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                return raw.decode(enc)
            except Exception:
                continue
        return raw.decode('utf-8', 'replace')

    @staticmethod
    def _peek_ooxml_kind(raw):
        """Detecta OOXML real por firma ZIP + rutas internas (ignora la extensión).

        Así un ``libro.xlsx`` renombrado a ``.xls`` (o un ``.docx`` a ``.doc``)
        sigue extrayéndose bien; el fallo típico era caer a decode texto binario.
        """
        if not raw or raw[:2] != b'PK':
            return None
        import io
        import zipfile
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                names = zf.namelist()
        except Exception:
            return None
        if any(n.startswith('xl/') for n in names):
            return 'xlsx'
        if 'word/document.xml' in names or any(n.startswith('word/') for n in names):
            return 'docx'
        if any(n.startswith('ppt/') for n in names):
            return 'pptx'
        return 'zip'

    @staticmethod
    def _is_ole_cfb(raw):
        """True si es Compound File Binary (``.xls`` / ``.doc`` antiguos)."""
        return bool(raw) and raw[:4] == b'\xD0\xCF\x11\xE0'

    @staticmethod
    def _format_excel_sheets(sheet_rows):
        """Une TODAS las pestañas con cabecera. Sin truncar filas ni hojas.

        ``sheet_rows``: lista de (title, list[str] filas tabuladas).
        """
        total = len(sheet_rows)
        titles = [t for t, _rows in sheet_rows]
        intro = (
            'Libro Excel: %d pestaña(s) — se listan TODAS (sin truncar): %s'
            % (total, ', '.join(u'«%s»' % t for t in titles))
        )
        blocks = [intro]
        for pos, (title, rows) in enumerate(sheet_rows, start=1):
            header = '===== HOJA %d/%d: «%s» (%d fila(s) con datos) =====' % (
                pos, total, title, len(rows),
            )
            body = '\n'.join(rows) if rows else '(hoja vacía)'
            blocks.append('%s\n%s' % (header, body))
        return '\n\n'.join(blocks)

    @classmethod
    def _xlsx_to_text(cls, raw):
        """Vuelca un .xlsx/.xlsm a texto con TODAS las pestañas (visibles y ocultas).

        Itera ``workbook.sheetnames`` (no solo la activa). Contenido completo,
        sin truncar.
        """
        import io
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                'openpyxl no está instalado en el servidor '
                '(pip install openpyxl; ya declarado en pns_ai_mcp)'
            )
        # read_only: streaming; data_only: valores calculados si existen en el fichero.
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            # sheetnames = TODAS las hojas del libro (incl. ocultas).
            names = list(wb.sheetnames or [])
            if not names:
                return 'Libro Excel: 0 pestañas (vacío)'
            sheet_rows = []
            for title in names:
                ws = wb[title]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = ['' if c is None else str(c) for c in row]
                    while cells and cells[-1] == '':
                        cells.pop()
                    if cells:
                        rows.append('\t'.join(cells))
                sheet_rows.append((title, rows))
            return cls._format_excel_sheets(sheet_rows)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @classmethod
    def _xls_to_text(cls, raw):
        """Vuelca un .xls (BIFF/OLE) a texto con TODAS las pestañas vía xlrd."""
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                'xlrd no está instalado en el servidor '
                '(pip install "xlrd>=1.2.0,<2" para .xls)'
            )
        book = xlrd.open_workbook(file_contents=raw, formatting_info=False)
        n = book.nsheets
        if n <= 0:
            return 'Libro Excel: 0 pestañas (vacío)'
        sheet_rows = []
        for idx in range(n):
            sh = book.sheet_by_index(idx)
            title = sh.name or ('Hoja%d' % (idx + 1))
            rows = []
            for r in range(sh.nrows):
                cells = []
                for c in range(sh.ncols):
                    val = sh.cell_value(r, c)
                    if val is None or val == '':
                        cells.append('')
                    else:
                        cells.append(str(val))
                while cells and cells[-1] == '':
                    cells.pop()
                if cells:
                    rows.append('\t'.join(cells))
            sheet_rows.append((title, rows))
        return cls._format_excel_sheets(sheet_rows)

    @staticmethod
    def _docx_to_text(raw):
        """Extrae texto de un .docx SIN dependencias: unzip + strip de tags XML."""
        import io
        import re
        import zipfile
        import html as _html
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            xml = z.read('word/document.xml').decode('utf-8', 'replace')
        # Saltos de párrafo/línea → \n antes de limpiar los tags.
        xml = re.sub(r'</w:p>', '\n', xml)
        xml = re.sub(r'<w:br\s*/?>', '\n', xml)
        text = re.sub(r'<[^>]+>', '', xml)
        return _html.unescape(text)

    def _extract_file_text(self, name, mimetype, b64):
        """Extrae el texto de un fichero de datos según su tipo. '' si no se puede."""
        if not b64:
            return ''
        try:
            raw = base64.b64decode(b64)
        except Exception:
            return ''
        lname = (name or '').lower()
        mt = (mimetype or '').lower()
        ooxml = self._peek_ooxml_kind(raw)
        is_ole = self._is_ole_cfb(raw)
        # Extensiones "cortas" .xls/.doc (no .xlsx/.docx)
        is_xls_ext = lname.endswith('.xls') and not lname.endswith('.xlsx')
        is_doc_ext = lname.endswith('.doc') and not lname.endswith('.docx')
        try:
            if lname.endswith('.pdf') or 'pdf' in mt:
                # Los PDF pasan por pns_ocr (texto embebido → visión si escaneado).
                # Es el gateway de "documento → texto" del suite (Fase 3).
                # pns_ocr es OPCIONAL: si no está instalado, se degrada con aviso.
                if 'ocr.service' not in self.env:
                    return '[PDF %s: OCR no instalado (módulo pns_ocr no disponible)]' % (name,)
                try:
                    return self.env['ocr.service'].extract_pdf(
                        raw, provider_id=self.provider_id or None,
                    )
                except Exception as _pdf_exc:
                    # Escaneado sin rasterizador u otro fallo: nota informativa
                    # (no rompe el turno) en vez de silencio.
                    _logger.warning(
                        'Chatboo async job %s: OCR PDF %s falló: %r',
                        self.id, name, _pdf_exc,
                    )
                    return '[PDF %s: no se pudo extraer texto (%s)]' % (
                        name, getattr(_pdf_exc, 'args', [''])[0] or type(_pdf_exc).__name__,
                    )
            # Excel OOXML (.xlsx / .xlsm) — por extensión, MIME o firma ZIP.
            # Siempre vuelca TODAS las pestañas (ver _xlsx_to_text).
            if (
                lname.endswith('.xlsx')
                or lname.endswith('.xlsm')
                or 'spreadsheetml' in mt
                or ooxml == 'xlsx'
            ):
                try:
                    return self._xlsx_to_text(raw)
                except ImportError as exc:
                    return '[Excel %s: %s]' % (name, exc)
                except Exception as exc:
                    _logger.warning(
                        'Chatboo async job %s: openpyxl falló en %s: %r',
                        self.id, name, exc,
                    )
                    return '[Excel %s: no se pudo leer (%s)]' % (
                        name, getattr(exc, 'args', [''])[0] or type(exc).__name__,
                    )
            # Excel BIFF (.xls real / OLE) — todas las pestañas vía xlrd.
            if is_ole and (is_xls_ext or 'ms-excel' in mt or mt == 'application/vnd.ms-excel'):
                try:
                    return self._xls_to_text(raw)
                except ImportError as exc:
                    return (
                        '[Excel %s: %s. Alternativa: guárdalo como .xlsx y adjúntalo.]'
                        % (name, exc)
                    )
                except Exception as exc:
                    _logger.warning(
                        'Chatboo async job %s: xlrd falló en %s: %r',
                        self.id, name, exc,
                    )
                    return '[Excel %s: no se pudo leer (.xls) (%s)]' % (
                        name, getattr(exc, 'args', [''])[0] or type(exc).__name__,
                    )
            # Word OOXML (.docx)
            if (
                lname.endswith('.docx')
                or 'wordprocessingml' in mt
                or ooxml == 'docx'
            ):
                try:
                    return self._docx_to_text(raw)
                except Exception as exc:
                    _logger.warning(
                        'Chatboo async job %s: docx falló en %s: %r',
                        self.id, name, exc,
                    )
                    return '[Word %s: no se pudo leer (%s)]' % (
                        name, getattr(exc, 'args', [''])[0] or type(exc).__name__,
                    )
            if is_ole and (is_doc_ext or 'msword' in mt or mt == 'application/msword'):
                return (
                    '[Word %s: formato .doc antiguo (binario). '
                    'Ábrelo en Word y guárdalo como .docx, luego adjúntalo de nuevo.]'
                    % (name,)
                )
            # Resto: texto plano / datos (txt, md, csv, tsv, log, json, ...).
            return self._decode_text_bytes(raw)
        except Exception:
            _logger.warning(
                'Chatboo async job %s: no se pudo extraer texto de %s (%s)',
                self.id, name, mimetype, exc_info=True,
            )
            return ''

    def _augment_prompt_with_files(self, base_message):
        """Devuelve el mensaje del turno con el texto de los ficheros adjuntos.

        No modifica self.prompt (lo que se muestra en la burbuja): el texto del
        fichero viaja SOLO al LLM en este turno. Sin truncar: el contenido
        extraído se inyecta entero (todas las pestañas Excel incluidas).
        """
        try:
            files = json.loads(self.files) if self.files else []
        except Exception:
            files = []
        if not files:
            return base_message
        blocks = []
        for f in files:
            name = (f or {}).get('name') or 'file'
            mimetype = (f or {}).get('mimetype') or ''
            _mt, b64 = self._split_data_url((f or {}).get('data') or '')
            text = self._extract_file_text(name, mimetype or _mt, b64)
            if not (text or '').strip():
                blocks.append('[Adjunto: %s] (no se pudo extraer texto)' % name)
                continue
            blocks.append('[Adjunto: %s]\n%s' % (name, text))
        if not blocks:
            return base_message
        joined = '\n\n'.join(blocks)
        if (base_message or '').strip():
            return '%s\n\n%s' % (base_message, joined)
        # Mensaje solo-fichero: instrucción mínima para el LLM.
        return 'Analiza el/los archivo(s) adjunto(s):\n\n%s' % joined

    def _persist_turn_files(self):
        """Persiste los ficheros del turno como ir.attachment y devuelve chips.

        Como las imágenes: el fichero se guarda una vez (ligado a la sesión) y el
        mensaje de usuario guarda un chip {name, url, mimetype} para mostrar un
        enlace descargable en la burbuja. El texto extraído NO se persiste.
        """
        import base64
        from odoo.addons.pns_ai_mcp.utils.session_download import (
            persist_chatboo_session_file,
        )
        try:
            files = json.loads(self.files) if self.files else []
        except Exception:
            files = []
        chips = []
        if not files:
            return chips
        for idx, f in enumerate(files):
            name = (f or {}).get('name') or ('archivo-%s' % (idx + 1))
            mimetype = (f or {}).get('mimetype') or ''
            _mt, b64 = self._split_data_url((f or {}).get('data') or '')
            if not b64:
                continue
            try:
                raw = base64.b64decode(b64)
                chip = persist_chatboo_session_file(
                    self.env, self.session_id.id, raw, name,
                    mimetype or _mt or 'application/octet-stream',
                )
                if chip:
                    chips.append({
                        'name': chip['name'],
                        'url': chip['url'],
                        'mimetype': chip['mimetype'],
                    })
            except Exception:
                _logger.warning(
                    'Chatboo async job %s: no se pudo persistir el fichero %s',
                    self.id, name, exc_info=True,
                )
        return chips

    def _llm_history_for_session(self, done_meta, visible):
        """Historial compacto para el siguiente turno (stubs, no HTML de burbuja).

        El motor lo emite en ``done.history``. Si falta (turno viejo / error),
        se reconstruye aquí para que el cliente no reenvíe tablas enormes.
        """
        meta = done_meta if isinstance(done_meta, dict) else {}
        hist = meta.get('history')
        if isinstance(hist, list) and hist:
            return hist
        try:
            from odoo.addons.pns_ai_mcp.utils.history_compact import (
                append_turn_stub,
                compact_history_for_llm,
            )
            try:
                incoming = json.loads(self.history) if self.history else []
            except Exception:
                incoming = []
            return append_turn_stub(
                compact_history_for_llm(incoming),
                self.prompt or '',
                visible=visible or '',
            )
        except Exception:
            _logger.debug(
                'Chatboo async job %s: could not compact LLM history',
                self.id, exc_info=True,
            )
            return None

    def _save_to_session(self, acc, done_meta, error, incomplete=False, not_started=False, query_code=None, query_data=None, active_skill=_SKILL_STATE_LEAVE, reclaim=False):
        """Añade el turno (usuario + asistente) al historial de la sesión.

        El contenido del asistente se guarda en crudo (markdown) con el flag
        ``raw``; el cliente lo formatea al recargar (reutiliza su formateador).

        ``incomplete=True`` marca una respuesta recuperada tras un reinicio: se
        guarda lo que se alcanzó a generar con una nota discreta y SIN estilo de
        error rojo (un reinicio no es un fallo del usuario). ``not_started=True``
        distingue el caso de un job que nunca llegó a arrancar (mensaje distinto).

        ``reclaim=True``: ruta ligera al cerrar un job zombie (reinicio Docker /
        worker muerto). No re-persiste base64 de imágenes/ficheros ni busca
        huérfanos: un prompt fallido no puede bloquear la apertura de Chatboo.
        """
        session = self.session_id
        if not session or not session.exists():
            return
        # Lectura fresca: el worker lleva su propia transacción abierta durante
        # toda la inferencia; invalidamos para no partir de un snapshot viejo.
        try:
            session.invalidate_recordset(['messages', 'staged_assistant_files'])
        except AttributeError:
            try:
                session.invalidate_cache(
                    ['messages', 'staged_assistant_files'], ids=session.ids,
                )
            except Exception:
                pass
        messages = session.get_messages()

        elapsed = None
        if self.started_at:
            elapsed = (fields.Datetime.now() - self.started_at).total_seconds()
        wall = self.env['chatboo.session'].format_display_datetime()
        stamp = wall
        if elapsed is not None:
            stamp = '%s (%ds)' % (wall, int(round(elapsed)))

        # Turno de usuario (para que quede persistido aunque el cliente no lo guarde).
        # Las imágenes pegadas se persisten como adjuntos y se referencian por URL
        # (NO se guarda base64 en el blob de la sesión: lo inflaría y se recarga
        # entero cada turno). Así el histórico conserva las capturas y son clicables.
        _user_msg = {
            'role': 'user',
            'content': self.prompt or '',
            'timestamp': wall,
        }
        _user_images = []
        _user_files = []
        if not reclaim:
            _user_images = self._persist_turn_images()
            if _user_images:
                _user_msg['images'] = _user_images
            _user_files = self._persist_turn_files()
            if _user_files:
                _user_msg['files'] = _user_files
        messages.append(_user_msg)

        _assistant_files = []
        _assistant_content = None
        _clip_data = None
        if not reclaim:
            try:
                from odoo.addons.pns_ai_mcp.utils.session_download import (
                    coalesce_download_chips,
                )
            except Exception:
                coalesce_download_chips = lambda *groups: [
                    c for g in groups for c in (g or [])
                    if isinstance(c, dict) and (c.get('url') or c.get('name'))
                ]
            _staged = session.consume_staged_assistant_download_chips()
            _orphans = session.recover_orphan_download_chips(
                messages=messages, since=self.started_at,
            )
            _assistant_files = coalesce_download_chips(
                (done_meta or {}).get('assistant_files'),
                _staged,
                _orphans,
            )
            _clip_data = (done_meta or {}).get('clip_data')
            svg_chips = []
            if (acc or '').strip() and not error:
                try:
                    from odoo.addons.pns_ai_mcp.utils.svg_download import (
                        persist_inline_svgs_from_html,
                    )
                    skill_code = None
                    if isinstance(active_skill, dict):
                        skill_code = active_skill.get('code') or None
                    if not skill_code:
                        skill_code = session.active_skill_code or None
                    svg_chips, acc = persist_inline_svgs_from_html(
                        self.env, session.id, acc,
                        prompt=self.prompt,
                        skill_code=skill_code,
                    )
                    svg_chips = svg_chips or []
                except Exception:
                    _logger.warning(
                        'Chatboo async job %s: no se pudieron persistir SVG inline',
                        self.id, exc_info=True,
                    )
            if svg_chips:
                _assistant_files = list(_assistant_files or []) + list(svg_chips)
                _assistant_content = acc or ''

        # Historial readline (flechas ↑/↓): persistir el prompt enviado para que
        # sobreviva al recargar la página. Con auto-promoción el cliente ya no
        # re-guarda la sesión (el worker autora el turno), así que si no lo
        # anotamos aquí el input_history se perdería. Dedup del último; el tope
        # (50) lo aplica set_input_history.
        if (self.prompt or '').strip():
            try:
                hist = session.get_input_history()
                if not hist or hist[-1] != self.prompt:
                    hist.append(self.prompt)
                    session.set_input_history(hist)
            except Exception:
                _logger.debug(
                    "Chatboo async: no se pudo actualizar input_history (sesión %s)",
                    session.id, exc_info=True,
                )

        if incomplete:
            # Recuperación tras reinicio: lo generado (si hay) + nota discreta.
            if (acc or '').strip():
                _note = _('(incomplete response: interrupted by a server restart)')
                _content = u"%s\n\n_%s_" % (acc, _note)
            elif not_started:
                _content = u"_%s_" % _(
                    'The request could not be started. Please resend.'
                )
            else:
                _content = u"_%s_" % _(
                    'The response was interrupted by a server restart. Please resend.'
                )
            messages.append({
                'role': 'assistant',
                'content': _content,
                'original_content': acc or '',
                'raw': True,
                'incomplete': True,
                'timestamp': stamp,
                **({'files': _assistant_files} if _assistant_files else {}),
            })
        elif error:
            messages.append({
                'role': 'assistant',
                'content': u"\U0001F6D1 %s\n\n%s" % (_('MCP engine error:'), error),
                'raw': True,
                'is_error': True,
                'timestamp': stamp,
                **({'files': _assistant_files} if _assistant_files else {}),
                **({'clip_data': _clip_data} if _clip_data else {}),
            })
        else:
            # Pie de respuesta: cada stack consume una forma distinta y hay que
            # reproducir AMBAS para que el pie sobreviva al recargar la sesión
            # (antes lo escribía el cliente en vivo; ahora lo autora el worker):
            #   · owl1 -> campos planos + model_details:{model,provider} + sources[]
            #   · owl2 -> objeto meta:{model,provider,context_limit,usage,sources,...}
            # Todo sale del evento 'done' del motor (done_meta), que ya trae
            # model, provider, protocol, context_limit, usage y sources.
            _dm = done_meta or {}
            _model = _dm.get('model') or ''
            _provider = _dm.get('provider') or ''
            _display_currency = _dm.get('display_currency') or ''
            _sources = _dm.get('sources') or []
            _records = _dm.get('records') or []
            _usage = _dm.get('usage') or None
            _ctx_limit = _dm.get('context_limit') or None
            _speed = _dm.get('speed_tps') or 0
            _pspeed = _dm.get('prompt_speed_tps') or 0
            _corr = (_dm.get('correlation_id') or '').strip()
            _user_prompt = (self.prompt or '').strip()
            _local = bool(_dm.get('local_ack'))
            if _local:
                _usage = None
                _ctx_limit = None
                _corr = ''
                _speed = 0
                _pspeed = 0
                _model = _model or 'Chatboo'
                _provider = _provider or 'local'
                for _prev in reversed(messages):
                    if isinstance(_prev, dict) and _prev.get('role') == 'user':
                        _prev['offtopic'] = True
                        break
            _model_details = (
                {
                    'model': _model,
                    'provider': _provider,
                    'display_currency': _display_currency,
                }
                if (_model or _provider) else None
            )
            _meta = {
                'model': _model,
                'provider': _provider,
                'display_currency': _display_currency,
                'protocol': _dm.get('protocol') or '',
                'context_limit': _ctx_limit,
                'usage': _usage,
                'sources': _sources,
                'records': _records,
                'speed_tps': _speed,
                'prompt_speed_tps': _pspeed,
                'correlation_id': _corr,
                'local_ack': _local,
                'user_prompt': _user_prompt,
            }
            messages.append({
                'role': 'assistant',
                'content': acc or '',
                'original_content': acc or '',
                'raw': True,
                # ── owl1 (campos planos que lee chatboo_component_v2.js) ──
                'usage': _usage,
                'context_limit': _ctx_limit,
                'speed_tps': _speed,
                'prompt_speed_tps': _pspeed,
                'correlation_id': _corr,
                'local_ack': _local,
                'user_prompt': _user_prompt,
                'offtopic': _local,
                'model_details': _model_details,
                'sources': _sources,
                'records': _records,
                'backend_history': self._llm_history_for_session(
                    _dm, acc,
                ),
                # ── owl2 (objeto meta que lee chatboo_app.js) ──
                'meta': _meta,
                'timestamp': stamp,
                **({'files': _assistant_files} if _assistant_files else {}),
                **({'clip_data': _clip_data} if _clip_data else {}),
            })

        session.set_messages(messages)
        # Persistir el ancla de reutilización solo cuando hay un código nuevo
        # exitoso (turno normal). En recuperaciones/errores llega None y se
        # conserva el de la última consulta buena.
        if query_code:
            try:
                session.write({'last_query_code': query_code})
            except Exception:
                _logger.debug(
                    "Chatboo async: no se pudo guardar last_query_code (sesión %s)",
                    session.id, exc_info=True,
                )
        # Nivel 2: cachear el dataset (filas) con su marca de tiempo para expirarlo
        # por antigüedad. query_data es None salvo en un turno normal con datos.
        if query_data is not None:
            try:
                session.write({
                    'last_query_data': json.dumps(query_data, default=str, ensure_ascii=False),
                    'last_query_data_date': fields.Datetime.now(),
                })
            except Exception:
                _logger.debug(
                    "Chatboo async: no se pudo guardar last_query_data (sesión %s)",
                    session.id, exc_info=True,
                )
        # Skill pegajoso: persistir el estado emitido por el motor. code con
        # valor → fijar/actualizar; code=None → soltar; sentinela → no tocar.
        if active_skill is not _SKILL_STATE_LEAVE:
            try:
                code = (active_skill or {}).get('code')
                if code:
                    session.write({
                        'active_skill_code': code,
                        'active_skill_params': json.dumps(
                            (active_skill or {}).get('params') or {},
                            default=str, ensure_ascii=False,
                        ),
                    })
                else:
                    session.write({
                        'active_skill_code': False,
                        'active_skill_params': False,
                    })
            except Exception:
                _logger.debug(
                    "Chatboo async: no se pudo guardar active_skill (sesión %s)",
                    session.id, exc_info=True,
                )
        session.update_last_used()
        # Devuelve los chips persistidos del turno de usuario (imágenes con URL
        # /web/image y ficheros con URL /web/content). El cliente los usa para
        # PARCHEAR en su sitio la burbuja recién enviada, que aún mostraba el
        # base64 (imágenes) y los ficheros sin URL (no clicables). Así se abren en
        # pestaña sin esperar a recargar la sesión.
        result = {
            'user_images': _user_images or [],
            'user_files': _user_files or [],
            'assistant_files': _assistant_files or [],
        }
        if _assistant_content is not None:
            result['assistant_content'] = _assistant_content
        return result

    # ──────────────────────────── Lectura (tail / poll) ────────────────────────────

    @api.model
    def read_progress(self, request_id):
        """Lectura de progreso en cursor propio (ve los commits del worker)."""
        try:
            with _get_registry(self.env.cr.dbname).cursor() as cr:
                cr.execute(
                    "SELECT state, partial, struct_events, done_meta, response, error "
                    "FROM chatboo_async_request WHERE id=%s",
                    (int(request_id),),
                )
                row = cr.fetchone()
        except Exception:
            return None
        if not row:
            return None
        state, partial, struct_events, done_meta, response, error = row
        return {
            'state': state,
            'partial': partial or '',
            'struct_events': json.loads(struct_events) if struct_events else [],
            'done_meta': json.loads(done_meta) if done_meta else {},
            'response': response or '',
            'error': error or '',
        }

    def mark_seen(self):
        """Marca el job como entregado a un cliente (evita re-entrega por poll)."""
        for rec in self:
            try:
                with _get_registry(rec.env.cr.dbname).cursor() as cr:
                    cr.execute(
                        "UPDATE chatboo_async_request SET seen=true WHERE id=%s", (rec.id,),
                    )
            except Exception:
                _logger.debug("Chatboo async mark_seen failed (request %s)", rec.id, exc_info=True)

    def request_cancel(self):
        """Marca el job para cancelación cooperativa (SSE abort + resume F5).

        Si aún está ``pending``, se cierra al momento. Si ``running``, el worker
        corta en el siguiente evento / flush y finaliza con aviso.
        """
        self.ensure_one()
        if self.user_id.id != self.env.uid:
            return False
        if self.state not in ('pending', 'running'):
            return False
        self.write({'cancel_requested': True})
        if self.state == 'pending':
            msg = '⏹ ' + _('Generation cancelled.')
            self.write({
                'state': 'done',
                'response': msg,
                'partial': msg,
                'finished_at': fields.Datetime.now(),
                'seen': False,
            })
            try:
                self._save_to_session(msg, {}, error=None)
            except Exception:
                _logger.debug(
                    "Chatboo async: cancel persist failed (job %s)",
                    self.id, exc_info=True,
                )
            try:
                self._notify('async_done')
            except Exception:
                pass
        return True

    def _is_cancel_requested(self):
        """Lee ``cancel_requested`` en cursor fresco (visible desde el worker)."""
        try:
            with _get_registry(self.env.cr.dbname).cursor() as cr:
                cr.execute(
                    "SELECT cancel_requested FROM chatboo_async_request WHERE id=%s",
                    (self.id,),
                )
                row = cr.fetchone()
                return bool(row and row[0])
        except Exception:
            return False

    # ──────────────────────────── Mantenimiento (cron) ────────────────────────────

    @api.model
    def _reclaim_stuck(self, minutes=None):
        """Recupera los jobs colgados (worker muerto tras un reinicio).

        Gracias al latido (``write_date`` se refresca en cada flush), solo caen
        aquí los jobs REALMENTE muertos. En vez de marcarlos como error rojo,
        RECUPERAMOS lo que se alcanzó a generar (``partial``) y lo guardamos en la
        sesión como respuesta incompleta, cerrando el job como 'done' y avisando
        como info (no como fallo). Así la campanita/respuesta sobrevive al reinicio.

        ``minutes`` permite un umbral más corto para el reclaim inmediato al abrir
        Chatboo (ver ``async_poll``); por defecto usa ``STUCK_MINUTES`` (cron).

        SEGURIDAD: gracias al latido independiente (``_start_heartbeat``), un job
        vivo refresca ``write_date`` cada ~10s, así que ``write_date < cutoff`` solo
        cae en jobs REALMENTE muertos. No matamos nunca un job que sigue latiendo,
        por lento que sea el modelo.

        NO BLOQUEANTE: ``FOR UPDATE SKIP LOCKED`` + savepoint por job + ruta
        ``reclaim=True`` (sin re-persistir base64). Un job zombie / prompt fallido
        NUNCA puede colgar la apertura de Chatboo ni el cron de mantenimiento.
        """
        mins = STUCK_MINUTES if minutes is None else minutes
        cutoff = fields.Datetime.now() - timedelta(minutes=mins)
        timeout_ms = int(self.env.context.get('chatboo_reclaim_timeout_ms', 8000) or 8000)
        cutoff_s = fields.Datetime.to_string(cutoff)
        # Candidatos sin lock; cada uno se reclama con SKIP LOCKED (no esperar).
        try:
            candidates = self.search([
                ('state', 'in', ('pending', 'running')),
                ('write_date', '<', cutoff),
            ]).ids
        except Exception:
            _logger.exception("Chatboo async: reclaim candidate search failed")
            return

        for rid in candidates:
            try:
                self.env.cr.execute(
                    """
                    SELECT id FROM chatboo_async_request
                     WHERE id = %s
                       AND state IN ('pending', 'running')
                       AND write_date < %s
                       FOR UPDATE SKIP LOCKED
                    """,
                    (rid, cutoff_s),
                )
                if not self.env.cr.fetchone():
                    continue
            except Exception:
                _logger.exception(
                    "Chatboo async: reclaim lock failed for job %s", rid,
                )
                continue

            try:
                with self.env.cr.savepoint():
                    job = self.browse(rid)
                    if not job.exists() or job.state not in ('pending', 'running'):
                        continue
                    partial = job.partial or ''
                    never_started = job.state == 'pending' and not partial
                    _logger.warning(
                        "Chatboo async: reclaiming stuck job %s (%s), recovering %s chars",
                        job.id, job.state, len(partial),
                    )
                    # Cerrar YA el estado dentro del savepoint; si el save a
                    # sesión falla/timeout, el fallback SQL deja el job done.
                    job.write({
                        'state': 'done',
                        'response': partial,
                        'error': False,
                        'finished_at': fields.Datetime.now(),
                    })
                    if timeout_ms > 0:
                        try:
                            self.env.cr.execute(
                                "SET LOCAL statement_timeout = %s",
                                (int(timeout_ms),),
                            )
                        except Exception:
                            pass
                    try:
                        job._save_to_session(
                            partial, {}, error=None, incomplete=True,
                            not_started=never_started, reclaim=True,
                        )
                    finally:
                        if timeout_ms > 0:
                            try:
                                self.env.cr.execute(
                                    "SET LOCAL statement_timeout = DEFAULT",
                                )
                            except Exception:
                                pass
                    try:
                        job._notify('async_done')
                    except Exception:
                        pass
            except Exception:
                _logger.exception("Chatboo async: reclaim failed for job %s", rid)
                # Último recurso: SQL crudo en savepoint limpio para no dejar
                # el zombie en running ni abortar la TX del caller (poll/cron).
                try:
                    with self.env.cr.savepoint():
                        self.env.cr.execute(
                            """
                            UPDATE chatboo_async_request
                               SET state='done',
                                   error=%s,
                                   finished_at=(now() at time zone 'UTC')
                             WHERE id=%s
                               AND state IN ('pending', 'running')
                            """,
                            (
                                'reclaim failed after server restart; please resend',
                                rid,
                            ),
                        )
                except Exception:
                    _logger.exception(
                        "Chatboo async: raw reclaim mark failed (job %s)", rid,
                    )
    @api.model
    def _gc_old(self):
        """Purga jobs terminados antiguos."""
        cutoff = fields.Datetime.now() - timedelta(hours=GC_HOURS)
        old = self.search([
            ('state', 'in', ('done', 'error')),
            ('finished_at', '<', cutoff),
        ])
        if old:
            old.unlink()

    @api.model
    def cron_maintenance(self):
        """Punto de entrada del cron: reclama colgados y purga viejos."""
        try:
            self._reclaim_stuck()
        except Exception:
            _logger.exception("Chatboo async: cron reclaim failed")
        # Liberar locks antes del GC: un reclaim lento no debe retener el
        # ir.cron "busy" mientras purga (otros workers lo saltaban eternamente).
        if not _is_test_mode(self.env):
            try:
                self.env.cr.commit()
            except Exception:
                _logger.debug("Chatboo async: cron commit after reclaim failed", exc_info=True)
        try:
            self._gc_old()
        except Exception:
            _logger.exception("Chatboo async: cron GC failed")
        # Nivel 2: purgar datasets cacheados obsoletos (por antigüedad) para que
        # no se acumulen en sesiones de larga vida. No debe tumbar el cron.
        try:
            self.env['chatboo.session']._gc_stale_query_data()
        except Exception:
            _logger.debug("Chatboo async: _gc_stale_query_data falló", exc_info=True)

    # ──────────────────────────── Bus ────────────────────────────

    def _notify(self, action):
        """Avisa al usuario por el bus. Compatible Odoo 14 y 17+.

        El front (componente + systray) ya entiende ``pns_chatboo_sync`` con
        ``action`` in {thinking, async_done, error}.
        """
        self.ensure_one()
        partner = self.user_id.partner_id
        if not partner:
            return
        payload = {
            'type': 'pns_chatboo_sync',
            'action': action,
            'session_id': self.session_id.id,
            'request_id': self.id,
            'is_error': action == 'error',
        }
        bus = self.env['bus.bus']
        try:
            # Odoo 16/17+: _sendone(target, notification_type, message)
            bus._sendone(partner, 'pns_chatboo_sync', payload)
            return
        except (AttributeError, TypeError):
            pass
        try:
            # Odoo 14: sendone(channel, message)
            channel = (self.env.cr.dbname, 'res.partner', partner.id)
            bus.sendone(channel, json.dumps(payload, ensure_ascii=False))
        except Exception:
            _logger.debug("Chatboo async bus notify failed (request %s)", self.id, exc_info=True)
